from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import Usuario, Impresora, PC, Notebook, Telefono, AltaImpresora, AltaTelefono, AltaPc, AltaNotebook, ActivoInfraestructura, AltaActivo, Empresa, AbonoCelular, AbonoImpresora, Licencia, AltaLicenciaoffice, AbonoOffice
from .form import UsuariosForm, ImpresorasForm, PcsForm, TelefonosForm, NotebooksForm, AltaImpresoraForm, AltaTelefonoForm, AltaPcForm, AltaNotebookForm, ActivosForm, AltaActivoForm, AbonoCelForm, AbonoImpForm, LicenciaForm, AbonoOffForm
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.urls import reverse
from .models import EliminacionImpresora, EliminacionPc, EliminacionNotebook, EliminacionTelefono, EliminacionActivo
from django.views.decorators.http import require_GET
from django.contrib import messages
from django.db.models import Sum
from .form import EliminacionImpresoraForm, EliminacionTelefonoForm, EliminacionNotebookForm, EliminacionPcForm, EliminacionActivoForm
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook

# Create your views here.

#INICIAR SESION
def iniciar(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('inicio')  # Redirigir al usuario a la URL con nombre "inicio"
        else:
            error = 'El usuario o la password son incorrectos.'
            return render(request, 'iniciar.html', {'error': error})
    else:
        return render(request, 'iniciar.html')

# INICIO
@login_required
def inicio(request):

    # USUARIOS DESCONOCIDOS

    cantidad_teams_rooms_pro = Licencia.objects.filter(licenciaoffice='microsoftteamsroomspro').count()
    cantidad_office_e5 = Licencia.objects.filter(licenciaoffice='office365E5').count()
    cantidad_office_e5_wac = Licencia.objects.filter(licenciaoffice='office365E5WAC').count()
    cantidad_exchange_plan_1 = Licencia.objects.filter(licenciaoffice='exchangeonlineplan1').count()
    cantidad_exchange_kiosk = Licencia.objects.filter(licenciaoffice='exchangeonlinekiosk').count()
    cantidad_office_e1 = Licencia.objects.filter(licenciaoffice='office365e1').count()
    
    usuarios = Usuario.objects.none()
    usuarios2 = Usuario.objects.all()
    empresas = Empresa.objects.all()
    
    telefonos = Telefono.objects.all().count
    impresoras = Impresora.objects.all().count
    pcs = PC.objects.all().count
    notebooks = Notebook.objects.all().count
    licencias = Licencia.objects.all().count()
    
    # TOTAL STOCK ASIGNADOS
    
    telefonos_count = Telefono.objects.exclude(usuario_id=611).count()
    notebooks_count = Notebook.objects.exclude(usuario_id=611).count()
    pcs_count = PC.objects.exclude(usuario_id=611).count()
    impresoras_count = Impresora.objects.exclude(usuario_id=611).count()
    activos_count = ActivoInfraestructura.objects.exclude(usuario_id=611).count()
    
    # TOTAL SIN ASIGNAR
    
    telefonos_sin_asignar = Telefono.objects.filter(usuario_id=611).count()
    notebooks_sin_asignar = Notebook.objects.filter(usuario_id=611).count()
    pcs_sin_asignar = PC.objects.filter(usuario_id=611).count()
    impresoras_sin_asignar = Impresora.objects.filter(usuario_id=611).count()
    activos_sin_asignar = ActivoInfraestructura.objects.filter(usuario_id=611).count()
    
    #TOTAL DE IRREPARABLES

    telefonos_irre = Telefono.objects.filter(reparabilidad='Irreparable').count()
    pcs_irre = PC.objects.filter(reparabilidad='Irreparable').count()
    impresoras_irre = Impresora.objects.filter(reparabilidad='Irreparable').count()
    notebooks_irre = Notebook.objects.filter(reparabilidad='Irreparable').count()
        
    total_irre = telefonos_irre + pcs_irre + impresoras_irre + notebooks_irre
    print(total_irre)
     
    buscar_query = request.GET.get('buscar1')
    
    # Obtener todos los usuarios
    usuarios = Usuario.objects.all()
    
    usuarioscont = usuarios.count()

    if buscar_query:
        # Dividir la consulta en palabras clave individuales
        palabras_clave = buscar_query.split()

        # Construir una consulta compleja utilizando Q objects
        consulta = Q()
        for palabra in palabras_clave:
            consulta |= Q(nombre_apellido__icontains=palabra)
            consulta |= Q(area__icontains=palabra)
            consulta |= Q(cargo__icontains=palabra)
            consulta |= Q(legajo__icontains=palabra)
            consulta |= Q(CATEGORIA_CTT__icontains=palabra)
            consulta |= Q(CCO__icontains=palabra)
            consulta |= Q(DNI__icontains=palabra)
            consulta |= Q(lugar_trab__icontains=palabra)
            consulta |= Q(lab_lpg__icontains=palabra)

        # Filtrar los usuarios basados en la consulta construida
        usuarios = usuarios.filter(consulta)

    # PAGINACION DE USUARIOS
    paginator_usuarios = Paginator(usuarios, 1000)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    # CANTIDAD TOTAL DE USUARIOS
    total_usuarios = usuarios.count()
    
    if request.method == 'GET':
        
        marca = request.GET.get('marca', '')

        context = {

            'usuarios': usuarios2,
            'licencias':licencias,
            'cantidad_teams_rooms_pro': cantidad_teams_rooms_pro,
            'cantidad_office_e5': cantidad_office_e5,
            'cantidad_office_e5_wac': cantidad_office_e5_wac,
            'cantidad_exchange_plan_1': cantidad_exchange_plan_1,
            'cantidad_exchange_kiosk': cantidad_exchange_kiosk,
            'cantidad_office_e1': cantidad_office_e1,
            'usuario': usuarios.first() if usuarios else None,
            'telefonos':telefonos,
            'impresoras':impresoras,
            'pcs':pcs,
            'notebooks':notebooks,
            'marcas': Empresa.MARCA_CHOICES,
            'telefonos_count': telefonos_count,
            'notebooks_count': notebooks_count,
            'pcs_count': pcs_count,
            'impresoras_count': impresoras_count,
            'activos_count':activos_count,
            'total_irre':total_irre,
            'telefonos_sin_asignar':telefonos_sin_asignar,
            'notebooks_sin_asignar':notebooks_sin_asignar,
            'pcs_sin_asignar':pcs_sin_asignar,
            'impresoras_sin_asignar':impresoras_sin_asignar,
            'activos_sin_asignar':activos_sin_asignar,
            'page_obj_usuarios': page_obj_usuarios,
            'usuarios': usuarios, 'page_obj_usuarios': page_obj_usuarios, 'total_usuarios': total_usuarios
        }
        
        
        if marca:

            marca_slug = marca.replace(" ","")
            return redirect(f'marca/{marca_slug}')
       

        return render(request, 'inicio.html', context)

# REGISTRO
def registro(request):
    if request.method =="GET" :

        return render (request, 'registro.html',{
            'form': UserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(
                    username=request.POST['username'],
                    password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('inicio')
            except IntegrityError:
                return render (request, 'registro.html',{
                   'form': UserCreationForm,
                   "error": 'El usuario ya existe'
                 })

        return render (request, 'registro.html',{
                   'form': UserCreationForm,
                   "error": 'Las contrasenas no coinciden'
                 })


#ABM
@login_required
def abm(request):

    #USUARIOS
    usuarios = Usuario.objects.all()

    #PAGINACION USUARIO

    paginator_usuarios = Paginator(usuarios,2)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    #CANTIDAD TOTAL DE USUARIOS
    total_usuarios = Usuario.objects.count()

    return render(request, 'sistema/abm.html',{'total_usuarios':total_usuarios, 'page_obj_usuarios':page_obj_usuarios})

#STOCK TOTAL
@login_required
def stock_total(request):
    # USUARIOS
    usuarios = Usuario.objects.all()

    # PAGINACION USUARIO
    paginator_usuarios = Paginator(usuarios, 2)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    # CANTIDAD TOTAL DE USUARIOS
    total_usuarios = Usuario.objects.count()

    # TOTAL STOCK
    telefonos_count = Telefono.objects.count()
    notebooks_count = Notebook.objects.count()
    pcs_count = PC.objects.count()
    impresoras_count = Impresora.objects.count()
    activos_count = ActivoInfraestructura.objects.count()
    
    
    #TOTAL DE IRREPARABLES
    telefonos_irre = Telefono.objects.filter(reparabilidad='Irreparable').count()
    pcs_irre = PC.objects.filter(reparabilidad='Irreparable').count()
    impresoras_irre = Impresora.objects.filter(reparabilidad='Irreparable').count()
    notebooks_irre = Notebook.objects.filter(reparabilidad='Irreparable').count()
    
    total_irre = telefonos_irre + pcs_irre + impresoras_irre + notebooks_irre
    print(total_irre)
    context = {
        'telefonos_count': telefonos_count,
        'notebooks_count': notebooks_count,
        'pcs_count': pcs_count,
        'impresoras_count': impresoras_count,
        'activos_count':activos_count
    }

    return render(request, 'sistema/stock_total.html', {'total_usuarios': total_usuarios, 'page_obj_usuarios': page_obj_usuarios,'total_irre':total_irre, **context})

#CERRAR SESION
@login_required
def cerrar(request):
    logout(request)
    return redirect('iniciar')

#USUARIO 
@login_required
def usuarios_crear(request):
    if request.method == 'POST':
        form = UsuariosForm(request.POST)
        if form.is_valid():
            usuarios = form.save(commit=False)
            usuarios.save()
            print(usuarios)
            return redirect('inicio')
    else:
        form = UsuariosForm()
    return render(request, 'sistema/usuarios_crear.html', {'form': form})

#EDITAR ACTIVOS
@login_required
def activos_usuario_id(request, usuario_id):
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    pcs = PC.objects.filter(usuario=usuario)
    
    telefonos = Telefono.objects.filter(usuario=usuario)
    
    impresoras = Impresora.objects.filter(usuario=usuario)
    
    notebooks = Notebook.objects.filter(usuario=usuario)


    totalpc = pcs.count()
    totaltel = telefonos.count()
    totalimp = impresoras.count()
    totalnote = notebooks.count()
    
    

    if request.method == 'POST':
        # Procesar el formulario enviado por POST aqui
        # Actualizar los activos del usuario segun los datos recibidos

        # Redirigir a una pagina de exito o a la pagina de detalles del usuario actualizado
        return redirect('usuarios_listar', usuario_id=usuario_id)

    return render(request, 'sistema/activos_usuario_id.html', {'usuario': usuario, 'pcs': pcs, 'telefonos': telefonos, 'impresoras': impresoras, 'notebooks': notebooks,'totalpc':totalpc,'totaltel':totaltel,'totalimp':totalimp,'totalnote':totalnote})

#MOSTRAR_TODO
@login_required
def usuarios_listar(request):
    buscar_query = request.GET.get('buscar1')
    
    # Obtener todos los usuarios
    usuarios = Usuario.objects.all()
    
    usuarioscont = usuarios.count()

    if buscar_query:
        # Dividir la consulta en palabras clave individuales
        palabras_clave = buscar_query.split()

        # Construir una consulta compleja utilizando Q objects
        consulta = Q()
        for palabra in palabras_clave:
            consulta |= Q(nombre_apellido__icontains=palabra)
            consulta |= Q(area__icontains=palabra)
            consulta |= Q(cargo__icontains=palabra)
            consulta |= Q(legajo__icontains=palabra)
            consulta |= Q(CATEGORIA_CTT__icontains=palabra)
            consulta |= Q(CCO__icontains=palabra)
            consulta |= Q(DNI__icontains=palabra)
            consulta |= Q(lugar_trab__icontains=palabra)
            consulta |= Q(lab_lpg__icontains=palabra)

        # Filtrar los usuarios basados en la consulta construida
        usuarios = usuarios.filter(consulta)

    # PAGINACION DE USUARIOS
    paginator_usuarios = Paginator(usuarios, 20)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    # CANTIDAD TOTAL DE USUARIOS
    total_usuarios = usuarios.count()

    return render(request, 'sistema/usuarios_listar.html', {'usuarios': usuarios, 'page_obj_usuarios': page_obj_usuarios, 'total_usuarios': total_usuarios})

# EDITAR USUARIOS
@login_required
def usuarios_editar(request, usuario_id):
    usuarios = get_object_or_404(Usuario, id=usuario_id)
    form = UsuariosForm(instance=usuarios)  # Inicializar el objeto form aqui
    if request.method == 'POST':
        form = UsuariosForm(request.POST, instance=usuarios)
        if form.is_valid():
            form.save()
            return redirect('usuarios_listar')
    return render(request, 'sistema/usuarios_editar.html',{'form': form, 'usuarios': usuarios})

# ELIMINAR USUARIOS
@login_required
def usuarios_eliminar(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)

    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'El usuario ha sido eliminado correctamente.')
        return redirect('usuarios_listar')

    return render(request, 'sistema/usuarios_eliminar.html', {'usuario': usuario})




#USUARIOS_DATOS
@login_required
def usuarios_datos(request, usuario_id):
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    

    return render(request, 'sistema/usuarios_datos.html',{'usuario': usuario})

#ACTIVOS USUARIO
@login_required
def activos_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)

    telefonos = Telefono.objects.filter(usuario=usuario)
    notebooks = Notebook.objects.filter(usuario=usuario)
    impresoras = Impresora.objects.filter(usuario=usuario)
    pcs = PC.objects.filter(usuario=usuario)
    activosinfra = ActivoInfraestructura.objects.filter(usuario=usuario)
    licencia_usuario = Licencia.objects.filter(usuario=usuario)

    cant_tel = telefonos.count()
    cant_note = notebooks.count()
    cant_imp = impresoras.count()
    cant_pcs = pcs.count()
    cant_activosinfra = activosinfra.count()
    cant_licencia = licencia_usuario.count()

    activos = {
        'telefonos': telefonos,
        'notebooks': notebooks,
        'impresoras': impresoras,
        'pcs': pcs,
        'activosinfra':activosinfra,
        'cant_tel': cant_tel,
        'licencia_usuario': licencia_usuario,
        'cant_licencia':cant_licencia,
    }

    return render(request, 'sistema/activos_usuario.html', {'activos': activos, 'usuario': usuario, 'telefonos':telefonos, 'notebooks':notebooks, 'pcs':pcs,'impresoras':impresoras, 'activosinfra':activosinfra, 'cant_tel':cant_tel, 'cant_note':cant_note, 'cant_imp':cant_imp, 'cant_pcs':cant_pcs, 'cant_activosinfra':cant_activosinfra,'licencia_usuario':licencia_usuario,'cant_licencia':cant_licencia})

""" #EMPRESAS
@login_required
def empresas_crear(request):
    if request.method == 'POST':
        form = EmpresasForm(request.POST)
        if form.is_valid():
            empresas = form.save(commit=False)
            empresas.save()
            print(empresas)
            return redirect('inicio')
    else:
        form = EmpresasForm()
    return render(request, 'sistema/empresas_crear.html', {'form': form})

@login_required
def empresas_listar(request):
    empresa_id = request.GET.get('empresa_id')
    area_select = request.GET.get('area', '')

    empresa_seleccionada = get_object_or_404(Empresa, id=empresa_id)

    if area_select:
        usuarios = Usuario.objects.filter(area=area_select)
        telefonos = Telefono.objects.filter(usuario__in=usuarios, empresa=empresa_seleccionada)
        notebooks = Notebook.objects.filter(usuario__in=usuarios, empresa=empresa_seleccionada)
        impresoras = Impresora.objects.filter(usuario__in=usuarios, empresa=empresa_seleccionada)
        pcs = PC.objects.filter(usuario__in=usuarios, empresa=empresa_seleccionada)
    else:
        telefonos = Telefono.objects.filter(empresa=empresa_seleccionada)
        notebooks = Notebook.objects.filter(empresa=empresa_seleccionada)
        impresoras = Impresora.objects.filter(empresa=empresa_seleccionada)
        pcs = PC.objects.filter(empresa=empresa_seleccionada)
        
    
        totaltel = telefonos.count()
        totalimp = impresoras.count()
        totalnote = notebooks.count()
        totalpcs = pcs.count()
  
    context = {
        'page_obj_usuarios': telefonos,
        'page_obj_usuarios1': notebooks,
        'page_obj_usuarios2': impresoras,
        'page_obj_usuarios3': pcs,
        'areas_todas': Usuario.AREAS_CHOICES,
        'area_select': area_select,
        'empresa_seleccionada': empresa_seleccionada,
        'totaltel' : totaltel,
        'totalimp':totalimp,
        'totalnote':totalnote,
        'totalpcs':totalpcs
    }
    

    return render(request, 'sistema/empresas_listar.html',context)

@login_required
def empresas_eliminar(request, empresa_id):
    empresa = get_object_or_404(Empresa, id=empresa_id)
    if request.method == 'POST':
        empresa.delete()
        return redirect('empresas_listar')
    return render(request, 'sistema/empresas_eliminar.html',{'empresa':empresa})

@login_required
def empresas_editar(request, empresa_id):
    empresas = get_object_or_404(Empresa, id=empresa_id)
    form = EmpresasForm(instance=empresas)  # Inicializar el objeto form aqui
    if request.method == 'POST':
        form = EmpresasForm(request.POST, instance=empresas)
        if form.is_valid():
            form.save()
            return redirect('empresas_listar')
    return render(request, 'sistema/empresas_editar.html',{'form': form, 'empresas': empresas}) """

#IMPRESORAS
@login_required
def impresoras_crear(request):
    if request.method == 'POST':
        
        form = ImpresorasForm(request.POST)
        alta_form = AltaImpresoraForm(request.POST)
        
        if form.is_valid() and alta_form.is_valid:
            impresora = form.save(commit=False)
            impresora.save()
            
            fecha_alta= timezone.now()
            usuario_alta = request.user
            alta = AltaImpresora.objects.create(
                usuario_alta = usuario_alta,
                fecha_alta=fecha_alta,
                impresora=impresora
            )
            
            return redirect('inicio')
    else:
        form = ImpresorasForm()
        alta_form = AltaImpresora()
        
    # Filtra los usuarios segun tus criterios y pasalos al contexto del renderizado
    usuarios = Usuario.objects.all()
    form.fields['usuario'].queryset = usuarios  # Establecer el queryset del campo usuario

    return render(request, 'sistema/impresoras_crear.html', {'form': form, 'alta_form':alta_form})


#ACTIVOS
@login_required
def activos_listar(request):

    contador = ActivoInfraestructura.objects.count()
    buscar_query = request.GET.get('buscaract')
    otro_filtro = request.GET.get('otrofiltro')
    
    if buscar_query:
        if buscar_query == 'AVENUE CORDOBA':
            usuarios = Usuario.objects.filter(Q(lab_lpg='AVENUE CORDOBA') | Q(lab_lpg='ADMINISTRACION CENTRAL') | Q(lab_lpg='PEUGEOT CORDOBA')).exclude(lugar_trab='SANTA FE')
            activos = ActivoInfraestructura.objects.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(reparabilidad__icontains=buscar_query) | Q(tipo__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query) | Q(labplg__icontains=buscar_query) | Q(area__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
            contador = activos.count()
        else:
            usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=buscar_query) | Q(cargo__icontains=buscar_query) | Q(area__icontains=buscar_query)
            )
            activos = ActivoInfraestructura.objects.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(reparabilidad__icontains=buscar_query) | Q(tipo__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query) | Q(labplg__icontains=buscar_query) | Q(area__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        contador = activos.count()
    else:
        activos = ActivoInfraestructura.objects.exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
    if otro_filtro:
        usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=otro_filtro) | Q(cargo__icontains=otro_filtro) | Q(area__icontains=otro_filtro)
            )
        activos = activos.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=otro_filtro) | Q(reparabilidad__icontains=otro_filtro) | Q(tipo__icontains=otro_filtro) | Q(modelo__icontains=otro_filtro) | Q(sn__icontains=otro_filtro) | Q(labplg__icontains=otro_filtro) | Q(area__icontains=otro_filtro)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        contador = activos.count()
        
    paginator_activos = Paginator(activos, 500)
    page_number_activos = request.GET.get('page')
    page_obj_activos = paginator_activos.get_page(page_number_activos)
    
    return render(request, 'sistema/activos_listar.html', {'activos': activos, 'page_obj_activos': page_obj_activos, 'contador':contador})

@login_required
def impresoras_listar(request):
    
    contador = Impresora.objects.count()
    buscar_query = request.GET.get('buscarimp')
    otro_filtro = request.GET.get('otrofiltro')
    tipo = request.GET.get('tipo')

    impresoras = Impresora.objects.exclude(reparabilidad='Irreparable')
    
    if tipo:
        impresoras = impresoras.filter(tipo=tipo)
        contador = impresoras.count()


    if buscar_query:
        if buscar_query == 'AVENUE CORDOBA':
            usuarios = Usuario.objects.filter(Q(lab_lpg='AVENUE CORDOBA') | Q(lab_lpg='ADMINISTRACION CENTRAL') | Q(lab_lpg='PEUGEOT CORDOBA')).exclude(lugar_trab='SANTA FE')
            impresoras = impresoras.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(abono__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query)
            ).order_by('usuario__nombre_apellido')
            contador = impresoras.count()
        else:
            usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=buscar_query) | Q(lab_lpg__icontains=buscar_query) | Q(cargo__icontains=buscar_query)
            )
            impresoras = impresoras.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(abono__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query)
            ).order_by('usuario__nombre_apellido')
            contador = impresoras.count()
    
    if otro_filtro:
        usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=otro_filtro) | Q(lab_lpg__icontains=otro_filtro) | Q(cargo__icontains=otro_filtro) | Q(area__icontains=otro_filtro)
            )
        impresoras = impresoras.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=otro_filtro) | Q(reparabilidad__icontains=otro_filtro) | Q(abono__icontains=otro_filtro) | Q(modelo__icontains=otro_filtro) | Q(sn__icontains=otro_filtro)
            ).order_by('usuario__nombre_apellido')
        contador = impresoras.count()
    
    paginator_impresoras = Paginator(impresoras, 500)
    page_number_impresoras = request.GET.get('page')
    page_obj_impresoras = paginator_impresoras.get_page(page_number_impresoras)
    
    return render(request, 'sistema/impresoras_listar.html', {'impresoras': impresoras, 'page_obj_impresoras': page_obj_impresoras, 'contador':contador, 'buscar_query':buscar_query, 'tipo':tipo})

#DESCARGAR EXCEL IMP
def descargar_excel_imp(request):
    contador = Impresora.objects.count()
    buscar_query = request.GET.get('buscarimp')
    otro_filtro = request.GET.get('otrofiltro')
    
    if buscar_query:
        if buscar_query == 'PEUGEOT CORDOBA':
            usuarios = Usuario.objects.filter(Q(lab_lpg='AVENUE CORDOBA') | Q(lab_lpg='ADMINISTRACION CENTRAL')).exclude(lugar_trab='SANTA FE')
            impresoras = Impresora.objects.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(tipo__icontains=buscar_query) | Q(abono__icontains=buscar_query) | Q(modelo__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        else:
            usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=buscar_query) | Q(lab_lpg__icontains=buscar_query) | Q(cargo__icontains=buscar_query)
            )
            impresoras = Impresora.objects.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(tipo__icontains=buscar_query) | Q(abono__icontains=buscar_query) | Q(modelo__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
    
        contador = impresoras.count()
    else:
        impresoras = Impresora.objects.exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
    
    if otro_filtro:
        usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=otro_filtro) | Q(lab_lpg__icontains=otro_filtro) | Q(cargo__icontains=otro_filtro) | Q(area__icontains=otro_filtro)
            )
        impresoras = impresoras.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=otro_filtro) | Q(reparabilidad__icontains=otro_filtro) | Q(tipo__icontains=otro_filtro) | Q(modelo__icontains=otro_filtro)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    

    # Agregar encabezados
    headers = ['Usuario','Area', 'Estado', 'Tipo', 'Modelo', 'Abono', 'LabLPG']
    sheet.append(headers)

    # Agregar datos a las filas
    for impresora in impresoras:
        row = [
            impresora.usuario.nombre_apellido,
            impresora.usuario.area,
            impresora.estado,
            impresora.tipo,
            impresora.modelo,
            impresora.abono,
            impresora.lablpg,
        ]
        sheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=impresoras.xlsx'
    workbook.save(response)
    return response


#ELIMINAR ACTIVOS
@login_required
def activos_eliminar(request, activo_id):
    activo = get_object_or_404(ActivoInfraestructura, id=activo_id)
    usuario_activo = activo.usuario

    
    if request.method == 'POST':
        form = EliminacionActivoForm(request.POST)
        
        if form.is_valid():
            motivo = form.cleaned_data['motivo_eliminacion']
            fecha_elim = timezone.now()
            usuario_sin_asignar, _ = Usuario.objects.get_or_create(nombre_apellido='Sin_asignar')

            #Cambiar el usuario del activo a "Sin_asignar"
            activo.usuario = usuario_sin_asignar
            activo.save()
            
            #Crear instancia de EliminacionImpresora y asignar valores
            eliminacion = EliminacionActivo()
            eliminacion.activo = activo
            eliminacion.motivo_eliminacion = motivo
            eliminacion.usuario_eliminacion = request.user
            eliminacion.fecha_eliminacion = fecha_elim
            eliminacion.usuario_ant = usuario_activo.nombre_apellido
            eliminacion.save()
            
            #Redireccionar a la vista 'detalles.html'
            return redirect('detallesactivo', activo_id=activo.id)
    else:
        form = EliminacionActivoForm()
    
    return render(request, 'sistema/activos_eliminar.html' , {'activo':activo, 'form':form})


@login_required
def impresoras_eliminar(request, impresora_id):
    impresora = get_object_or_404(Impresora, id=impresora_id)

    usuario_impresora = impresora.usuario

    if request.method == 'POST':
        form = EliminacionImpresoraForm(request.POST)

        if form.is_valid():
            motivo = form.cleaned_data['motivo_eliminacion']
            fecha_elim = timezone.now()

            usuario_sin_asignar, _ = Usuario.objects.get_or_create(nombre_apellido='Sin_asignar')

            # Cambiar el usuario de la impresora a "Sin_asignar"
            impresora.usuario = usuario_sin_asignar
            impresora.save()

            # Crear instancia de EliminacionImpresora y asignar valores
            eliminacion = EliminacionImpresora()
            eliminacion.impresora = impresora
            eliminacion.motivo_eliminacion = motivo
            eliminacion.usuario_eliminacion = request.user  # Assign the logged-in user
            eliminacion.fecha_eliminacion = fecha_elim
            eliminacion.usuario_ant = usuario_impresora.nombre_apellido
            eliminacion.save()

            # Redireccionar a la vista 'detalles.html'
            return redirect('detalles', impresora_id=impresora.id)

    else:
        form = EliminacionImpresoraForm()

    return render(request, 'sistema/impresoras_eliminar.html', {'impresora': impresora, 'form': form})

@login_required
def activos_editar(request, activo_id):
    activo = get_object_or_404(ActivoInfraestructura, id=activo_id)
    form = ActivosForm(instance=activo)
    alta_form = AltaActivoForm()
    
    if request.method == 'POST':
        form = ActivosForm(request.POST, instance=activo)
        alta_form = AltaActivoForm(request.POST)
        
        if form.is_valid() and alta_form.is_valid():
            activo = form.save()
            
            fecha_alta = timezone.now()
            usuario_alta = request.user
            
            alta = AltaActivo.objects.create(
                activo=activo,
                usuario_alta=usuario_alta,
                fecha_alta=fecha_alta
            )
            return redirect('altasactivo', activo_id=activo_id)
    
    return render(request, 'sistema/activos_editar.html', {'form': form, 'alta_form': alta_form, 'activo':activo})

@login_required
def impresoras_editar(request, impresora_id):
    impresora = get_object_or_404(Impresora, id=impresora_id)
    form = ImpresorasForm(instance=impresora)
    alta_form = AltaImpresoraForm()  # Initialize AltaImpresoraForm
    
    if request.method == 'POST':
        form = ImpresorasForm(request.POST, instance=impresora)
        alta_form = AltaImpresoraForm(request.POST)  # Bind AltaImpresoraForm with POST data
        
        if form.is_valid() and alta_form.is_valid():
            impresora = form.save()
            
            fecha_alta = timezone.now()
            usuario_alta = request.user
            print(usuario_alta)
            alta = AltaImpresora.objects.create(
                impresora=impresora,
                usuario_alta=usuario_alta,
                fecha_alta=fecha_alta
            )

            return redirect('altas', impresora_id=impresora_id)

    return render(request, 'sistema/impresoras_editar.html', {'form': form, 'alta_form': alta_form, 'impresoras': impresora})

@login_required
def detalles(request, impresora_id):
    impresora = get_object_or_404(Impresora, id=impresora_id)
    eliminacion = EliminacionImpresora.objects.filter(impresora=impresora).first()
    return render(request, 'sistema/detalles.html', {'impresora': impresora, 'eliminacion': eliminacion})

@login_required
def detallespc(request, pc_id):
    pc = get_object_or_404(PC, id=pc_id)
    eliminacion = EliminacionPc.objects.filter(pc=pc).first()
    return render(request, 'sistema/detallespc.html', {'pc': pc, 'eliminacion': eliminacion})

@login_required
def detallesnote(request, notebook_id):
    notebook = get_object_or_404(Notebook, id=notebook_id)
    eliminacion = EliminacionNotebook.objects.filter(notebook=notebook).first()
    return render(request, 'sistema/detallesnote.html', {'notebook': notebook, 'eliminacion': eliminacion})

@login_required
def detallestel(request, telefono_id):
    telefono = get_object_or_404(Telefono, id=telefono_id)
    eliminacion = EliminacionTelefono.objects.filter(telefono=telefono).first()
    return render(request, 'sistema/detallestel.html', {'telefono': telefono, 'eliminacion': eliminacion})

@login_required
def detallesactivo(request, activo_id):
    activo = get_object_or_404(ActivoInfraestructura, id=activo_id)
    eliminacion = AltaActivo.objects.filter(activo=activo).first()
    return render(request, 'sistema/detallesactivo.html', {'activo': activo, 'eliminacion': eliminacion})


@login_required
def altas(request, impresora_id):
    impresora = get_object_or_404(Impresora, id=impresora_id)
    alta = AltaImpresora.objects.filter(impresora=impresora).order_by('-fecha_alta').first()
    usuario_alta = User.objects.get(id=alta.usuario_alta_id).username if alta else None
    return render(request, 'sistema/altas.html', {'impresora': impresora, 'alta': alta, 'usuario_alta': usuario_alta})

@login_required
def altasactivo(request, activo_id):
    activo = get_object_or_404(ActivoInfraestructura, id=activo_id)
    alta = AltaActivo.objects.filter(activo=activo).order_by('-fecha_alta').first()
    usuario_alta = User.objects.get(id=alta.usuario_alta_id).username if alta else None
    return render(request, 'sistema/altasactivo.html', {'activo': activo, 'alta': alta, 'usuario_alta': usuario_alta})

@login_required
def altastel(request, telefono_id):
    telefono = get_object_or_404(Telefono, id=telefono_id)
    alta = AltaTelefono.objects.filter(telefono=telefono).order_by('-fecha_alta').first()
    usuario_alta = User.objects.get(id=alta.usuario_alta_id).username if alta else None
    return render(request, 'sistema/altastel.html', {'telefono': telefono, 'alta': alta, 'usuario_alta': usuario_alta})

@login_required
def altaspc(request, pc_id):
    pc = get_object_or_404(PC, id=pc_id)
    alta = AltaPc.objects.filter(pc=pc).order_by('-fecha_alta').first()
    usuario_alta = User.objects.get(id=alta.usuario_alta_id).username if alta else None
    return render(request, 'sistema/altaspc.html', {'pc': pc, 'alta': alta, 'usuario_alta': usuario_alta})

@login_required
def altasnote(request, notebook_id):
    notebook = get_object_or_404(Notebook, id=notebook_id)
    alta = AltaNotebook.objects.filter(notebook=notebook).order_by('-fecha_alta').first()
    usuario_alta = User.objects.get(id=alta.usuario_alta_id).username if alta else None
    return render(request, 'sistema/altasnote.html', {'notebook': notebook, 'alta': alta, 'usuario_alta': usuario_alta})
    

@login_required
def lista_altas(request):
    todo = AltaImpresora.objects.all()
    return render(request, 'sistema/lista_altas.html', {'todo':todo})

@login_required
def irreparables(request):
    
    irreparable_impresoras = Impresora.objects.filter(reparabilidad='Irreparable')
    irreparable_telefonos = Telefono.objects.filter(reparabilidad='Irreparable')
    irreparable_notebooks = Notebook.objects.filter(reparabilidad='Irreparable')
    irreparable_pcs = PC.objects.filter(reparabilidad='Irreparable')
    irreparable_activos = ActivoInfraestructura.objects.filter(reparabilidad='Irreparable')


    cant_imp = irreparable_impresoras.count()
    cant_tel = irreparable_telefonos.count()
    cant_note = irreparable_notebooks.count()
    cant_pc = irreparable_pcs.count()
    cant_act = irreparable_activos.count()

    total = cant_imp + cant_tel + cant_note + cant_pc + cant_act

    return render(request, 'sistema/irreparables.html', {'total':total,'irreparable_activos':irreparable_activos,'cant_act':cant_act,'cant_pc':cant_pc,'cant_note':cant_note,'cant_tel':cant_tel,'cant_imp':cant_imp,'irreparable_impresoras': irreparable_impresoras,'irreparable_pcs':irreparable_pcs,'irreparable_notebooks':irreparable_notebooks,'irreparable_telefonos':irreparable_telefonos})

#PCS
@login_required
def pcs_crear(request):
    if request.method == 'POST':
        form = PcsForm(request.POST)
        alta_form = AltaPcForm(request.POST)
        if form.is_valid() and alta_form.is_valid:
            sn_valor = form.cleaned_data['sn']
            buscar_igual = PC.objects.filter(sn=sn_valor).exists()
            if not buscar_igual:
                pc = form.save(commit=False)
                pc.save()
                fecha_alta=timezone.now()
                usuario_alta = request.user
                alta = AltaPc.objects.create(
                    usuario_alta=usuario_alta,
                    fecha_alta=fecha_alta,
                    pc=pc
                )
                return redirect('inicio')
            else:
                form = PcsForm()
                alta_form = AltaPcForm()
                #Mensaje de duplicado
                messages.error(request, 'YA EXISTE UNA PC CON EL NUMERO DE SERIE INGRESADO.')
                return render(request, 'sistema/pcs_crear.html', {'form':form, 'alta_form':alta_form})
    else:
        form = PcsForm()
        alta_form= AltaPc()

    # Filtra los usuarios segun tus criterios y pasalos al contexto del renderizado
    usuarios = Usuario.objects.all()
    form.fields['usuario'].queryset = usuarios  # Establecer el queryset del campo usuario

    return render(request, 'sistema/pcs_crear.html', {'form': form, 'alta_form': alta_form})

@login_required
def pcs_listar(request):

    buscar_query = request.GET.get('buscarpc')
    otro_filtro = request.GET.get('otrofiltro')
    contador = PC.objects.count()

    if buscar_query:
        if buscar_query == 'AVENUE CORDOBA':
            usuarios = Usuario.objects.filter(Q(lab_lpg='AVENUE CORDOBA') | Q(lab_lpg='ADMINISTRACION CENTRAL') | Q(lab_lpg='PEUGEOT CORDOBA'))
            pcs = PC.objects.filter(
                Q(usuario__in=usuarios) | Q(procesador__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query) | Q(estado__icontains=buscar_query) | Q(monitor__icontains=buscar_query) | Q(disco__icontains=buscar_query) | Q(ram__icontains=buscar_query) | Q(reparabilidad__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        else:
            usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=buscar_query) | Q(lab_lpg__icontains=buscar_query) | Q(cargo__icontains=buscar_query) | Q(area__icontains=buscar_query)
            )
            pcs = PC.objects.filter(
                Q(usuario__in=usuarios) | Q(procesador__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query) | Q(estado__icontains=buscar_query) | Q(monitor__icontains=buscar_query) | Q(disco__icontains=buscar_query) | Q(ram__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

        contador = pcs.count()
    else:
        pcs = PC.objects.exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

    if otro_filtro:
        usuarios = Usuario.objects.filter(Q(nombre_apellido__icontains=otro_filtro) | Q(lab_lpg__icontains=otro_filtro) | Q(cargo__icontains=otro_filtro) | Q(area__icontains=otro_filtro))
        pcs = pcs.filter(
            Q(usuario__in=usuarios) | Q(procesador__icontains=otro_filtro) | Q(modelo__icontains=otro_filtro) | Q(sn__icontains=otro_filtro) | Q(estado__icontains=otro_filtro) | Q(monitor__icontains=otro_filtro) | Q(disco__icontains=otro_filtro) | Q(ram__icontains=otro_filtro)
        ).order_by('usuario__nombre_apellido')
        contador = pcs.count()

    paginator_pcs = Paginator(pcs, 500)
    page_number_pcs = request.GET.get('page')
    page_obj_pcs = paginator_pcs.get_page(page_number_pcs)

    return render(request, 'sistema/pcs_listar.html', {'pcs': pcs, 'page_obj_usuarios': page_obj_pcs, 'contador': contador})

@login_required
def pcs_eliminar(request, pc_id):
    pc = get_object_or_404(PC, id=pc_id)
    usuario_pc = pc.usuario
    
    if request.method == 'POST':
        form = EliminacionPcForm(request.POST)
        
        if form.is_valid():
            motivo = form.cleaned_data['motivo_eliminacion']
            fecha_elim = timezone.now()

            usuario_sin_asignar, _ = Usuario.objects.get_or_create(nombre_apellido='Sin_asignar')

            pc.usuario = usuario_sin_asignar
            pc.save()
            
            eliminacion = EliminacionPc()
            eliminacion.pc = pc
            eliminacion.motivo_eliminacion = motivo
            eliminacion.usuario_eliminacion = request.user
            eliminacion.fecha_eliminacion = fecha_elim
            eliminacion.usuario_ant = usuario_pc.nombre_apellido
            eliminacion.save()
            
            return redirect('detallespc', pc_id=pc.id)
    else:
        form = EliminacionPcForm()
    
    return render(request, 'sistema/pcs_eliminar.html', {'pc': pc, 'form': form})


@login_required
def pcs_editar(request, pc_id):
    pc = get_object_or_404(PC, id=pc_id)
    form = PcsForm(instance=pc)
    alta_form = AltaPcForm()

    if request.method == 'POST':
        form = PcsForm(request.POST, instance=pc)
        alta_form = AltaPcForm(request.POST)
        
        if form.is_valid() and alta_form.is_valid():
            pc = form.save()
            
            fecha_alta = timezone.now()
            usuario_alta = request.user
            alta = AltaPc.objects.create(
                pc=pc,
                usuario_alta=usuario_alta,
                fecha_alta=fecha_alta
            )
            
            return redirect('altaspc', pc_id=pc_id)
        
    return render(request, 'sistema/pcs_editar.html', {'form': form, 'pc': pc, 'alta_form': alta_form})

#TELEFONOS
@login_required
def telefonos_crear(request):
    if request.method == 'POST':
        form = TelefonosForm(request.POST)
        alta_form = AltaTelefonoForm(request.POST)
        if form.is_valid() and alta_form.is_valid():
            imei_o_sn_valor = form.cleaned_data['imei_o_sn']
            buscar_igual = Telefono.objects.filter(imei_o_sn=imei_o_sn_valor).exists()
            if not buscar_igual:
                telefono = form.save(commit=False)
                telefono.save()
                fecha_alta = timezone.now()
                usuario_alta = request.user
                alta = AltaTelefono.objects.create(
                    usuario_alta=usuario_alta,
                    fecha_alta=fecha_alta,
                    telefono=telefono
                )
                return redirect('inicio')
            else:
                form = TelefonosForm()
                alta_form = AltaTelefonoForm()
                # Mensaje de duplicado
                messages.error(request, 'YA EXISTE UN TELEFONO CON EL NUMERO DE SERIE INGRESADO, INTENTE CON OTRO DIFERENTE.')
                return render(request, 'sistema/telefonos_crear.html', {'form': form, 'alta_form': alta_form})
    else:
        form = TelefonosForm()
        alta_form = AltaTelefonoForm()
    
    usuarios = Usuario.objects.all()
    form.fields['usuario'].queryset = usuarios
    return render(request, 'sistema/telefonos_crear.html', {'form': form, 'alta_form': alta_form})

@login_required
def telefonos_listar(request):

    telefonos = Telefono.objects.exclude(usuario__nombre_apellido='Sin_asignar')
    
    contador = telefonos.count()
    otro_filtro = request.GET.get('otrofiltro')
    buscar_query = request.GET.get('buscartel')
    
    if buscar_query:
        if buscar_query == 'AVENUE CORDOBA':
            lablpg_values = ['AVENUE CORDOBA', 'PEUGEOT CORDOBA', 'ADMINISTRACION CENTRAL']
            telefonos = Telefono.objects.filter(
                (Q(estado__icontains=buscar_query) | Q(numero__icontains=buscar_query) | Q(imei_o_sn__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(marca__icontains=buscar_query) | Q(accesorio__icontains=buscar_query) | Q(plan__icontains=buscar_query) | Q(empresa_abono__icontains=buscar_query) | Q(rs__icontains=buscar_query) | Q(lablpg__in=lablpg_values))
                & ~Q(usuario__nombre_apellido='Sin_asignar')
             ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        elif buscar_query == 'VOLANT URQUIZA':
             usuarios = Usuario.objects.filter(Q(lab_lpg='VOLANT URQUIZA'))
             telefonos = Telefono.objects.filter(
                 (Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(numero__icontains=buscar_query) | Q(imei_o_sn__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(marca__icontains=buscar_query) | Q(accesorio__icontains=buscar_query) | Q(plan__icontains=buscar_query) | Q(empresa_abono__icontains=buscar_query) | Q(rs__icontains=buscar_query) | Q(lablpg__icontains=buscar_query))
                 & ~Q(usuario__nombre_apellido='Sin_asignar')
             ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        else:
            usuarios = Usuario.objects.filter(Q(nombre_apellido__icontains=buscar_query) | Q(cargo__icontains=buscar_query) | Q(area__icontains=buscar_query))
            telefonos = Telefono.objects.filter(
                (Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(numero__icontains=buscar_query) | Q(imei_o_sn__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(marca__icontains=buscar_query) | Q(accesorio__icontains=buscar_query) | Q(plan__icontains=buscar_query) | Q(empresa_abono__icontains=buscar_query) | Q(rs__icontains=buscar_query) | Q(lablpg__icontains=buscar_query))
                & ~Q(usuario__nombre_apellido='Sin_asignar')
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        contador = telefonos.count()
    else:
        telefonos = Telefono.objects.exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

    if otro_filtro:
        usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=otro_filtro) | Q(lab_lpg__icontains=otro_filtro) | Q(cargo__icontains=otro_filtro) | Q(area__icontains=otro_filtro)
            )
        telefonos = telefonos.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=otro_filtro) | Q(numero__icontains=otro_filtro) | Q(imei_o_sn__icontains=otro_filtro) | Q(modelo__icontains=otro_filtro) | Q(marca__icontains=otro_filtro) | Q(accesorio__icontains=otro_filtro) | Q(plan__icontains=otro_filtro) | Q(empresa_abono__icontains=otro_filtro) | Q(rs__icontains=otro_filtro) | Q(lablpg__icontains=otro_filtro)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        contador = telefonos.count()

    paginator_usuarios = Paginator(telefonos, 1000)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    return render(request, 'sistema/telefonos_listar.html', {'telefonos': telefonos, 'page_obj_usuarios': page_obj_usuarios, 'contador': contador,'buscar_query':buscar_query})

def descargar_excel(request):

    telefonos = Telefono.objects.exclude(usuario__nombre_apellido='Sin_asignar')

    contador = Telefono.objects.count()
    otro_filtro = request.GET.get('otrofiltro')
    buscar_query = request.GET.get('buscartel')
    
    if buscar_query:
        if buscar_query == 'PEUGEOT CORDOBA':
            lablpg_values = ['AVENUE CORDOBA', 'PEUGEOT CORDOBA', 'ADMINISTRACION CENTRAL']
            telefonos = Telefono.objects.filter(
                (Q(estado__icontains=buscar_query) | Q(numero__icontains=buscar_query) | Q(imei_o_sn__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(marca__icontains=buscar_query) | Q(accesorio__icontains=buscar_query) | Q(plan__icontains=buscar_query) | Q(empresa_abono__icontains=buscar_query) | Q(rs__icontains=buscar_query) | Q(lablpg__in=lablpg_values))
                & ~Q(usuario__nombre_apellido='Sin_asignar')
             ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        elif buscar_query == 'VOLANT URQUIZA':
             usuarios = Usuario.objects.filter(Q(lab_lpg='VOLANT URQUIZA'))
             telefonos = Telefono.objects.filter(
                 (Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(numero__icontains=buscar_query) | Q(imei_o_sn__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(marca__icontains=buscar_query) | Q(accesorio__icontains=buscar_query) | Q(plan__icontains=buscar_query) | Q(empresa_abono__icontains=buscar_query) | Q(rs__icontains=buscar_query) | Q(lablpg__icontains=buscar_query))
                 & ~Q(usuario__nombre_apellido='Sin_asignar')
             ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        else:
            usuarios = Usuario.objects.filter(Q(nombre_apellido__icontains=buscar_query) | Q(cargo__icontains=buscar_query) | Q(area__icontains=buscar_query))
            telefonos = Telefono.objects.filter(
                (Q(usuario__in=usuarios) | Q(estado__icontains=buscar_query) | Q(numero__icontains=buscar_query) | Q(imei_o_sn__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(marca__icontains=buscar_query) | Q(accesorio__icontains=buscar_query) | Q(plan__icontains=buscar_query) | Q(empresa_abono__icontains=buscar_query) | Q(rs__icontains=buscar_query) | Q(lablpg__icontains=buscar_query))
                & ~Q(usuario__nombre_apellido='Sin_asignar')
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        contador = telefonos.count()
    else:
        telefonos = Telefono.objects.exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

    if otro_filtro:
        usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=otro_filtro) | Q(lab_lpg__icontains=otro_filtro) | Q(cargo__icontains=otro_filtro) | Q(area__icontains=otro_filtro)
            )
        telefonos = telefonos.filter(
                Q(usuario__in=usuarios) | Q(estado__icontains=otro_filtro) | Q(numero__icontains=otro_filtro) | Q(imei_o_sn__icontains=otro_filtro) | Q(modelo__icontains=otro_filtro) | Q(marca__icontains=otro_filtro) | Q(accesorio__icontains=otro_filtro) | Q(plan__icontains=otro_filtro) | Q(empresa_abono__icontains=otro_filtro) | Q(rs__icontains=otro_filtro) | Q(lablpg__icontains=otro_filtro)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        contador = telefonos.count()
    
    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    

    # Agregar encabezados
    headers = ['Usuario','Area', 'Numero', 'Modelo', 'Marca', 'Accesorio', 'Plan', 'Empresa Abono', 'RS', 'LabLPG']
    sheet.append(headers)

    # Agregar datos a las filas
    for telefono in telefonos:
        row = [
            telefono.usuario.nombre_apellido,
            telefono.usuario.area,
            telefono.numero,
            telefono.modelo,
            telefono.marca,
            telefono.accesorio,
            telefono.plan,
            telefono.empresa_abono,
            telefono.rs,
            telefono.lablpg
        ]
        sheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=telefonos.xlsx'
    workbook.save(response)
    return response


@login_required
def telefonos_eliminar(request, telefono_id):

    telefono = get_object_or_404(Telefono, id=telefono_id)
    usuario_telefono = telefono.usuario
    
    if request.method == 'POST':
        form = EliminacionTelefonoForm(request.POST)  # Initialize the form
        
        if form.is_valid():
            motivo = form.cleaned_data['motivo_eliminacion']
            fecha_elim = timezone.now()

            usuario_sin_asignar, _ = Usuario.objects.get_or_create(nombre_apellido='Sin_asignar')
            
            telefono.usuario = usuario_sin_asignar
            telefono.save()
            
            eliminacion = EliminacionTelefono()
            eliminacion.telefono = telefono
            eliminacion.motivo_eliminacion = motivo
            eliminacion.usuario_eliminacion = request.user  # Assign the logged-in user
            eliminacion.fecha_eliminacion = fecha_elim
            eliminacion.usuario_ant = usuario_telefono.nombre_apellido
            eliminacion.save()
            
            return redirect('detallestel', telefono_id=telefono.id)
    else:
        form = EliminacionTelefonoForm()
    return render(request, 'sistema/telefonos_eliminar.html', {'telefono': telefono, 'form': form})

@login_required
def telefonos_editar(request, telefono_id):

    telefono = get_object_or_404(Telefono, id=telefono_id)
    form = TelefonosForm(instance=telefono)
    alta_form = AltaTelefonoForm()  # Inicia la variable del alta
      
    if request.method == 'POST':
        form = TelefonosForm(request.POST, instance=telefono)
        alta_form = AltaTelefonoForm(request.POST)
        
        if form.is_valid() and alta_form.is_valid():
            
            telefono = form.save()
            fecha_alta = timezone.now()
            usuario_alta = request.user
            print(usuario_alta)
            alta = AltaTelefono.objects.create(
                telefono=telefono,
                usuario_alta=usuario_alta,
                fecha_alta=fecha_alta
            )
            return redirect('altastel', telefono_id=telefono_id)

    return render(request, 'sistema/telefonos_editar.html', {'form': form, 'alta_form': alta_form, 'telefono': telefono})

#ACTIVOS_CREAR
@login_required
def activos_crear(request):
    if request.method == 'POST':
        form = ActivosForm(request.POST)
        alta_form = AltaActivoForm(request.POST)
        if form.is_valid() and alta_form.is_valid:
            sn_valor = form.cleaned_data['sn']
            buscar_igual = ActivoInfraestructura.objects.filter(sn=sn_valor).exists()
            
            if not buscar_igual:
                activo = form.save(commit=False)
                activo.save()
            
                fecha_alta=timezone.now()
                usuario_alta = request.user
                alta = AltaActivo.objects.create(
                    usuario_alta=usuario_alta,
                    fecha_alta=fecha_alta,
                    activo=activo
                )
                return redirect('inicio')
            else:
                form = ActivosForm()
                alta_form = AltaActivoForm()
                #Mensaje de duplicado
                messages.error(request, 'YA EXISTE UN ACTIVO CON EL NUMERO DE SERIE INGRESADO.')
                return render(request, 'sistema/telefonos_crear.html', {'form':form,'alta_form':alta_form})
    else:
        form = ActivosForm()
        alta_form= AltaActivoForm()
            
    usuarios = Usuario.objects.all()
    form.fields['usuario'].queryset = usuarios
    
    return render(request, 'sistema/activos_crear.html', {'form':form, 'alta_form':alta_form})

@login_required
def notebooks_crear(request):
    if request.method == 'POST':
        form = NotebooksForm(request.POST)
        alta_form = AltaNotebookForm(request.POST)

        if form.is_valid() and alta_form.is_valid():
            sn_valor = form.cleaned_data['sn']
            buscar_igual = Notebook.objects.filter(sn=sn_valor).exists()

            if not buscar_igual:
                notebook = form.save(commit=False)
                notebook.save()

                fecha_alta = timezone.now()
                usuario_alta = request.user
                alta = AltaNotebook.objects.create(
                    usuario_alta=usuario_alta,
                    fecha_alta=fecha_alta,
                    notebook=notebook
                )
                return redirect('inicio')
            else:
                form = NotebooksForm()
                alta_form = AltaNotebookForm()

                #Mensaje de duplicado
                messages.error(request, 'YA EXISTE UNA NOTEBOOK CON EL NUMERO DE SERIE INGRESADO.')

                return redirect('notebooks_crear')
    else:
        form = NotebooksForm()
        alta_form = AltaNotebookForm()

    usuarios = Usuario.objects.all()
    form.fields['usuario'].queryset = usuarios

    return render(request, 'sistema/notebooks_crear.html', {'form': form, 'alta_form': alta_form})

@login_required
def notebooks_listar(request):

    buscar_query = request.GET.get('buscarnote')
    otro_filtro = request.GET.get('otrofiltro')
    contador = Notebook.objects.count()

    if buscar_query:
        if buscar_query == 'AVENUE CORDOBA':
            usuarios = Usuario.objects.filter(Q(lab_lpg='AVENUE CORDOBA') | Q(lab_lpg='ADMINISTRACION CENTRAL') | Q(lab_lpg='PEUGEOT CORDOBA'))
            notebooks = Notebook.objects.filter(
                Q(usuario__in=usuarios) | Q(procesador__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query) | Q(estado__icontains=buscar_query) | Q(monitor__icontains=buscar_query) | Q(disco__icontains=buscar_query) | Q(ram__icontains=buscar_query) | Q(marca__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        else:
            usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=buscar_query) | Q(lab_lpg__icontains=buscar_query) | Q(cargo__icontains=buscar_query) | Q(area__icontains=buscar_query)
            )
            notebooks = Notebook.objects.filter(
                Q(usuario__in=usuarios) | Q(procesador__icontains=buscar_query) | Q(modelo__icontains=buscar_query) | Q(sn__icontains=buscar_query) | Q(estado__icontains=buscar_query) | Q(monitor__icontains=buscar_query) | Q(disco__icontains=buscar_query) | Q(ram__icontains=buscar_query) | Q(marca__icontains=buscar_query)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

        contador = notebooks.count()

    else:

        notebooks = Notebook.objects.exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

    if otro_filtro:
        usuarios = Usuario.objects.filter(
                Q(nombre_apellido__icontains=otro_filtro) | Q(lab_lpg__icontains=otro_filtro) | Q(cargo__icontains=otro_filtro) | Q(area__icontains=otro_filtro)
            )
        notebooks = notebooks.filter(
                Q(usuario__in=usuarios) | Q(marca__icontains=otro_filtro) | Q(procesador__icontains=otro_filtro) | Q(ram__icontains=otro_filtro) | Q(disco__icontains=otro_filtro) | Q(estado__icontains=otro_filtro) | Q(reparabilidad__icontains=otro_filtro) | Q(monitor__icontains=otro_filtro) | Q(modelo__icontains=otro_filtro) | Q(sn__icontains=otro_filtro) | Q(lablpg__icontains=otro_filtro)
            ).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        contador = notebooks.count()

    paginator_usuarios = Paginator(notebooks, 500)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    return render(request, 'sistema/notebooks_listar.html', {'notebooks': notebooks,'page_obj_usuarios':page_obj_usuarios,'contador':contador})

@login_required
def notebooks_eliminar(request, notebook_id):
    notebook = get_object_or_404(Notebook, id=notebook_id)

    usuario_notebook = notebook.usuario
    
    if request.method == 'POST':
        form = EliminacionNotebookForm(request.POST)

        if form.is_valid():
            motivo = form.cleaned_data['motivo_eliminacion']
            fecha_elim = timezone.now()
            usuario_sin_asignar, _ = Usuario.objects.get_or_create(nombre_apellido='Sin_asignar')

            # Cambiar el usuario del notebook a "Sin_asignar"
            notebook.usuario = usuario_sin_asignar
            notebook.save()

            # Crear instancia de EliminacionNotebook y asignar valores
            eliminacion = EliminacionNotebook()
            eliminacion.notebook = notebook
            eliminacion.motivo_eliminacion = motivo
            eliminacion.usuario_eliminacion = request.user  # Assign the logged-in user
            eliminacion.fecha_eliminacion = fecha_elim
            eliminacion.usuario_ant = usuario_notebook.nombre_apellido
            eliminacion.save()

            # Redireccionar a la vista 'detallesnote.html'
            return redirect('detallesnote', notebook_id=notebook.id)

    else:
        form = EliminacionNotebookForm()

    return render(request, 'sistema/notebooks_eliminar.html', {'notebook': notebook, 'form': form})


@login_required
def notebooks_editar(request, notebook_id):
    notebook = get_object_or_404(Notebook, id=notebook_id)
    form = NotebooksForm(instance=notebook)  # Inicializar el objeto form aqui
    alta_form = AltaNotebookForm()  # Initialize the alta_form variable here
    if request.method == 'POST':
        form = NotebooksForm(request.POST, instance=notebook)
        alta_form = AltaNotebookForm(request.POST)
        
        if form.is_valid() and alta_form.is_valid():
            notebook = form.save()
            
            fecha_alta = timezone.now()
            usuario_alta = request.user
            print(usuario_alta)
            alta = AltaNotebook.objects.create(
                notebook=notebook,
                usuario_alta=usuario_alta,
                fecha_alta=fecha_alta
            )
            return redirect('altasnote', notebook_id=notebook_id)
    return render(request, 'sistema/notebooks_editar.html',{'form': form, 'notebooks': notebook,'alta_form':alta_form})



#TOTAL STOCK
@login_required
def total_stock(request):
    
    notebooks = Notebook.objects.all()
    usuarios = Usuario.objects.all()
    impresoras = Impresora.objects.all()
    pcs = PC.objects.all()
    telefonos = Telefono.objects.all()
    
    return render(request, 'total_stock.html', {'notebooks': notebooks,'usuarios':usuarios,'impresoras':impresoras,'pcs':pcs,'telefonos':telefonos})


@login_required
def reportes(request):
    
    notebooks = Notebook.objects.all()
    usuarios = Usuario.objects.all()
    pcs = PC.objects.all()
    telefonos = Telefono.objects.all()
    impresoras = Impresora.objects.all()
    
    altanotebooks = AltaNotebook.objects.all()
    altatelefonos = AltaTelefono.objects.all()
    altapcs = AltaPc.objects.all()
    altaactivos = AltaActivo.objects.all()
    altaimpresoras = AltaImpresora.objects.all()

    total = ""
    totalnotes = ""
    totaltels = ""
    totalpcs = ""
    totalactivos = ""
    totalimps = ""

    users = User.objects.exclude(username__in=['sisemas', 'sistemas', 'mayramoises'])
    
    usuario_id = request.GET.get('usuario')
    usuario = User.objects.get(id=usuario_id) if usuario_id else None

    fecha = request.GET.get('fecha')
    print(fecha)

    if fecha:
        fecha_inicio = datetime.strptime(fecha, "%Y-%m-%d").date()
        fecha_fin = fecha_inicio + timedelta(days=1)

        totalnotes = altanotebooks.filter(fecha_alta__range=(fecha_inicio, fecha_fin)).count()
        totaltels = altatelefonos.filter(fecha_alta__range=(fecha_inicio, fecha_fin)).count()
        totalpcs = altapcs.filter(fecha_alta__range=(fecha_inicio, fecha_fin)).count()
        totalactivos = altaactivos.filter(fecha_alta__range=(fecha_inicio, fecha_fin)).count()
        totalimps = altaimpresoras.filter(fecha_alta__range=(fecha_inicio, fecha_fin)).count()

        altanotebooks = altanotebooks.filter(usuario_alta=usuario, fecha_alta__range=(fecha_inicio, fecha_fin))
        altatelefonos = altatelefonos.filter(usuario_alta=usuario, fecha_alta__range=(fecha_inicio, fecha_fin))
        altapcs = altapcs.filter(usuario_alta=usuario, fecha_alta__range=(fecha_inicio, fecha_fin))
        altaactivos = altaactivos.filter(usuario_alta=usuario, fecha_alta__range=(fecha_inicio, fecha_fin))
        altaimpresoras = altaimpresoras.filter(usuario_alta=usuario, fecha_alta__range=(fecha_inicio, fecha_fin))
    else:
        altanotebooks = altanotebooks.filter(usuario_alta=usuario)
        altatelefonos = altatelefonos.filter(usuario_alta=usuario)
        altapcs = altapcs.filter(usuario_alta=usuario)
        altaactivos = altaactivos.filter(usuario_alta=usuario)
        altaimpresoras = altaimpresoras.filter(usuario_alta=usuario)

        totalnotes = AltaNotebook.objects.all().count()
        totaltels = AltaTelefono.objects.all().count()
        totalpcs = AltaPc.objects.all().count()
        totalactivos = AltaActivo.objects.all().count()
        totalimps = AltaImpresora.objects.all().count()
        total = totalnotes + totaltels + totalpcs + totalactivos + totalimps 


    
    paginator_usuarios = Paginator(altanotebooks, 1000)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)
    
    paginator_usuarios = Paginator(altatelefonos, 1000)
    page_number_usuarios = request.GET.get('page')
    page_obj_telefonos = paginator_usuarios.get_page(page_number_usuarios)
    
    paginator_usuarios = Paginator(altapcs, 1000)
    page_number_usuarios = request.GET.get('page')
    page_obj_pcs = paginator_usuarios.get_page(page_number_usuarios)
    
    paginator_usuarios = Paginator(altaactivos, 1000)
    page_number_usuarios = request.GET.get('page')
    page_obj_activos = paginator_usuarios.get_page(page_number_usuarios)
    
    paginator_usuarios = Paginator(altaimpresoras, 1000)
    page_number_usuarios = request.GET.get('page')
    page_obj_impresoras = paginator_usuarios.get_page(page_number_usuarios)

    cantidad_notebooks = altanotebooks.filter(usuario_alta=usuario).count()
    cantidad_telefonos = altatelefonos.filter(usuario_alta=usuario).count()
    cantidad_pcs = altapcs.filter(usuario_alta=usuario).count()
    cantidad_impresoras = altaimpresoras.filter(usuario_alta=usuario).count()
    cantidad_activos = altaactivos.filter(usuario_alta=usuario).count()

    cantidad_total_individual = cantidad_notebooks + cantidad_telefonos + cantidad_pcs + cantidad_impresoras + cantidad_activos
    
    return render(request, 'reportes.html', {'altanotebooks': altanotebooks, 'page_obj_usuarios': page_obj_usuarios, 'users': users, 'usuario': usuario, 'altatelefonos':altatelefonos, 'page_obj_telefonos':page_obj_telefonos, 'altapcs':altapcs, 'page_obj_pcs':page_obj_pcs, 'altaactivo':altaactivos, 'page_obj_activos':page_obj_activos, 'altaimpresoras':altaimpresoras, 'page_obj_impresoras':page_obj_impresoras, 'cantidad_notebooks':cantidad_notebooks, 'cantidad_telefonos':cantidad_telefonos,'cantidad_pcs':cantidad_pcs,'cantidad_impresoras':cantidad_impresoras,'cantidad_activos':cantidad_activos,'totalnotes':totalnotes, 'totaltels':totaltels, 'totalpcs':totalpcs, 'totalactivos':totalactivos, 'totalimps':totalimps, 'cantidad_total_individual':cantidad_total_individual, 'total':total})





    
@login_required
def reportes_total(request):
    
    usuarios = Usuario.objects.none()
    usuarios2 = Usuario.objects.all()
    empresas = Empresa.objects.all()

    telefonos = Telefono.objects.all().count
    impresoras = Impresora.objects.all().count
    pcs = PC.objects.all().count
    notebooks = Notebook.objects.all().count
    

    context = {
            'usuarios': usuarios2,
            'usuario': usuarios.first() if usuarios else None,
            'telefonos':telefonos,
            'impresoras':impresoras,
            'pcs':pcs,
            'notebooks':notebooks,
            
        }

    return render(request, 'reportes_total.html', context)



#ELEGIR
@login_required
def elegir(request, usuario_id):
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    return render(request, 'elegir.html', {'usuario': usuario})

#SIN ASIGNAR
@login_required
def registros_sin_asignar(request):

    usuario_sin_asignar = Usuario.objects.filter(nombre_apellido='Sin_asignar').first()
    id_usuario_sin_asignar = usuario_sin_asignar.id

    telefonos_sin_asignar = Telefono.objects.filter(usuario_id=id_usuario_sin_asignar)
    impresoras_sin_asignar = Impresora.objects.filter(usuario_id=id_usuario_sin_asignar)
    pcs_sin_asignar = PC.objects.filter(usuario_id=id_usuario_sin_asignar)
    notebooks_sin_asignar = Notebook.objects.filter(usuario_id=id_usuario_sin_asignar)
    activos_sin_asignar = ActivoInfraestructura.objects.filter(usuario_id=id_usuario_sin_asignar)
    
    totaltel = telefonos_sin_asignar.count()
    totalimp = impresoras_sin_asignar.count()
    totalnote = notebooks_sin_asignar.count()
    totalpcs = pcs_sin_asignar.count()
    totalactivos = activos_sin_asignar.count()
    
    context = {
        
        'usuario': usuario_sin_asignar,
        'id_usuario': id_usuario_sin_asignar,
        'telefonos': telefonos_sin_asignar,
        'impresoras': impresoras_sin_asignar,
        'pcs': pcs_sin_asignar,
        'activosinfra': activos_sin_asignar,
        'notebooks': notebooks_sin_asignar,
        'totaltel' : totaltel,
        'totalimp':totalimp,
        'totalnote':totalnote,
        'totalpcs':totalpcs,
        'totalactivos':totalactivos,

    }

    return render(request, 'sistema/sin_asignar.html', context)



#BUSQUEDA POR USUARIO
""" def busqueda_por_usuario(request):
    usuario_id = request.GET.get('usuario_id')
    usuario = get_object_or_404(Usuario, id=usuario_id)
    telefonos = usuario.telefono_set.all()

    return render(request, 'sistema/busqueda_por_usuario.html', {'usuario': usuario, 'telefonos': telefonos}) """

""" #BUSQEUDA POR EMPRESA
def busqueda_por_empresa(request):
    empresa_id = request.GET.get('empresa_id')
    empresa = get_object_or_404(Empresa, id=empresa_id)
    
    return render(request, 'sistema/busqueda_por_empresa.html', {'empresa': empresa}) """

#BUSQUEDA POR AREA
@login_required
def busqueda_por_area(request):
    return render(request, 'sistema/busqueda_por_area.html')


#AREA
#REPUESTOS
@login_required
def repuestos(request, company=None):

    usuarios_repuestos = Usuario.objects.filter(area='REPUESTOS')
    
    usuarios_repuestos = usuarios_repuestos.filter(area='REPUESTOS', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_repuestos = Usuario.objects.filter(area='REPUESTOS')
            usuarios_repuestos = usuarios_repuestos.filter(area='REPUESTOS', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_repuestos = Usuario.objects.filter(area='REPUESTOS')
            usuarios_repuestos = usuarios_repuestos.filter(area='REPUESTOS', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_repuestos = Usuario.objects.filter(area='REPUESTOS')
            usuarios_repuestos = usuarios_repuestos.filter(area='REPUESTOS', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_repuestos = Usuario.objects.filter(area='REPUESTOS')
            usuarios_repuestos = usuarios_repuestos.filter(area='REPUESTOS', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_repuestos = Usuario.objects.filter(area='REPUESTOS')
            usuarios_repuestos = usuarios_repuestos.filter(area='REPUESTOS', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    
    telefonos_repuestos = []
    for usuario in usuarios_repuestos:
        telefonos_usuario_repuestos = usuario.telefono_set.all()
        telefonos_repuestos.extend(telefonos_usuario_repuestos)

    impresoras_repuestos = []
    for usuario1 in usuarios_repuestos:
        impresoras_usuario_repuestos = usuario1.impresora_set.all()
        impresoras_repuestos.extend(impresoras_usuario_repuestos)

    notebooks_repuestos = []
    for usuario2 in usuarios_repuestos:
        notebooks_usuario_repuestos = usuario2.notebook_set.all()
        notebooks_repuestos.extend(notebooks_usuario_repuestos)

    pcs_repuestos = []
    for usuario3 in usuarios_repuestos:
        pcs_usuario_repuestos = usuario3.pc_set.all()
        pcs_repuestos.extend(pcs_usuario_repuestos)

    activos_repuestos = []
    for usuario4 in usuarios_repuestos:
        activos_usuario_repuestos = usuario4.activoinfraestructura_set.all()
        activos_repuestos.extend(activos_usuario_repuestos)
            
    pctotal = len(pcs_repuestos)
    teltotal = len(telefonos_repuestos)
    imptotal = len(impresoras_repuestos)
    notetotal = len(notebooks_repuestos)
    acttotal = len(activos_repuestos)

    total = pctotal + teltotal + imptotal + notetotal + acttotal
    
    paginator = Paginator(telefonos_repuestos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_repuestos, 20)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_repuestos, 20)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_repuestos, 20)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_repuestos, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/REPUESTOS.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#VENTA DIRECTA 0 KM SUC
@login_required
def ventadirecta0kmsuc(request, company=None):

    usuarios_ventadirecta0kmsuc = Usuario.objects.filter(area='VENTA DIRECTA 0KM SUC')
    
    if company:
        usuarios_ventadirecta0kmsuc = usuarios_ventadirecta0kmsuc.filter(area='VENTA DIRECTA 0KM SUC', lab_lpg=company)

    telefonos_ventadirecta0kmsuc = []
    for usuario in usuarios_ventadirecta0kmsuc:
        telefonos_usuario_ventadirecta0kmsuc = usuario.telefono_set.all()
        telefonos_ventadirecta0kmsuc.extend(telefonos_usuario_ventadirecta0kmsuc)

    impresoras_ventadirecta0kmsuc = []
    for usuario1 in usuarios_ventadirecta0kmsuc:
        impresoras_usuario_ventadirecta0kmsuc = usuario1.impresora_set.all()
        impresoras_ventadirecta0kmsuc.extend(impresoras_usuario_ventadirecta0kmsuc)

    notebooks_ventadirecta0kmsuc = []
    for usuario2 in usuarios_ventadirecta0kmsuc:
        notebooks_usuario_ventadirecta0kmsuc = usuario2.notebook_set.all()
        notebooks_ventadirecta0kmsuc.extend(notebooks_usuario_ventadirecta0kmsuc)

    pcs_ventadirecta0kmsuc = []
    for usuario3 in usuarios_ventadirecta0kmsuc:
        pcs_usuario_ventadirecta0kmsuc = usuario3.pc_set.all()
        pcs_ventadirecta0kmsuc.extend(pcs_usuario_ventadirecta0kmsuc)


    paginator = Paginator(telefonos_ventadirecta0kmsuc, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_ventadirecta0kmsuc, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_ventadirecta0kmsuc, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_ventadirecta0kmsuc, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render(request, 'area/VENTADIRECTA0KMSUC.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3})

#ADM VENTAS
@login_required
def admventas(request, company=None):

    usuarios_admventas = Usuario.objects.filter(area='ADM VENTAS')
    
    usuarios_admventas = usuarios_admventas.filter(area='ADM VENTAS', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_admventas = Usuario.objects.filter(area='ADM VENTAS')
            usuarios_admventas = usuarios_admventas.filter(area='ADM VENTAS', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_admventas = Usuario.objects.filter(area='ADM VENTAS')
            usuarios_admventas = usuarios_admventas.filter(area='ADM VENTAS', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_admventas = Usuario.objects.filter(area='ADM VENTAS')
            usuarios_admventas = usuarios_admventas.filter(area='ADM VENTAS', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_admventas = Usuario.objects.filter(area='ADM VENTAS')
            usuarios_admventas = usuarios_admventas.filter(area='ADM VENTAS', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_admventas = Usuario.objects.filter(area='ADM VENTAS')
            usuarios_admventas = usuarios_admventas.filter(area='ADM VENTAS', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    telefonos_admventas = []
    for usuario in usuarios_admventas:
        telefonos_usuario_admventas = usuario.telefono_set.all()
        telefonos_admventas.extend(telefonos_usuario_admventas)

    impresoras_admventas = []
    for usuario1 in usuarios_admventas:
        impresoras_usuario_admventas = usuario1.impresora_set.all()
        impresoras_admventas.extend(impresoras_usuario_admventas)

    notebooks_admventas = []
    for usuario2 in usuarios_admventas:
        notebooks_usuario_admventas = usuario2.notebook_set.all()
        notebooks_admventas.extend(notebooks_usuario_admventas)

    pcs_admventas = []
    for usuario3 in usuarios_admventas:
        pcs_usuario_admventas = usuario3.pc_set.all()
        pcs_admventas.extend(pcs_usuario_admventas)
        
    activos_admventas = []
    for usuario4 in usuarios_admventas:
        activos_usuario_admventas = usuario4.activoinfraestructura_set.all()
        activos_admventas.extend(activos_usuario_admventas)
        
    pctotal = len(pcs_admventas)
    teltotal = len(telefonos_admventas)
    imptotal = len(impresoras_admventas)
    notetotal = len(notebooks_admventas)
    acttotal = len(activos_admventas)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_admventas, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_admventas, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_admventas, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_admventas, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_admventas, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/ADMVENTAS.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#SIN CLASIFICAR
@login_required
def sinclasificar(request, company=None):

    usuarios_sinclasificar = Usuario.objects.filter(area='SIN CLASIFICAR')
    
    usuarios_sinclasificar = usuarios_sinclasificar.filter(area='SIN CLASIFICAR', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_sinclasificar = Usuario.objects.filter(area='SIN CLASIFICAR')
            usuarios_sinclasificar = usuarios_sinclasificar.filter(area='SIN CLASIFICAR', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_sinclasificar = Usuario.objects.filter(area='SIN CLASIFICAR')
            usuarios_sinclasificar = usuarios_sinclasificar.filter(area='SIN CLASIFICAR', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_sinclasificar = Usuario.objects.filter(area='SIN CLASIFICAR')
            usuarios_sinclasificar = usuarios_sinclasificar.filter(area='SIN CLASIFICAR', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_sinclasificar = Usuario.objects.filter(area='SIN CLASIFICAR')
            usuarios_sinclasificar = usuarios_sinclasificar.filter(area='SIN CLASIFICAR', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_sinclasificar = Usuario.objects.filter(area='SIN CLASIFICAR')
            usuarios_sinclasificar = usuarios_sinclasificar.filter(area='SIN CLASIFICAR', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    telefonos_sinclasificar = []
    for usuario in usuarios_sinclasificar:
        telefonos_usuario_sinclasificar = usuario.telefono_set.all()
        telefonos_sinclasificar.extend(telefonos_usuario_sinclasificar)

    impresoras_sinclasificar = []
    for usuario1 in usuarios_sinclasificar:
        impresoras_usuario_sinclasificar = usuario1.impresora_set.all()
        impresoras_sinclasificar.extend(impresoras_usuario_sinclasificar)

    notebooks_sinclasificar = []
    for usuario2 in usuarios_sinclasificar:
        notebooks_usuario_sinclasificar = usuario2.notebook_set.all()
        notebooks_sinclasificar.extend(notebooks_usuario_sinclasificar)

    pcs_sinclasificar = []
    for usuario3 in usuarios_sinclasificar:
        pcs_usuario_sinclasificar = usuario3.pc_set.all()
        pcs_sinclasificar.extend(pcs_usuario_sinclasificar)

    activos_sinclasificar = []
    for usuario4 in usuarios_sinclasificar:
        activos_usuario_sinclasificar = usuario4.activoinfraestructura_set.all()
        activos_sinclasificar.extend(activos_usuario_sinclasificar)

    pctotal = len(pcs_sinclasificar)
    teltotal = len(telefonos_sinclasificar)
    imptotal = len(impresoras_sinclasificar)
    notetotal = len(notebooks_sinclasificar)
    acttotal = len(activos_sinclasificar)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_sinclasificar, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_sinclasificar, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_sinclasificar, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_sinclasificar, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_sinclasificar, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/SINCLASIFICAR.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#AMD POST VENTA
@login_required
def admpostventa(request, company=None):

    usuarios_admpostventa = Usuario.objects.filter(area='ADM POST VENTA')
    
    usuarios_admpostventa = usuarios_admpostventa.filter(area='ADM POST VENTA', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_admpostventa = Usuario.objects.filter(area='ADM POST VENTA')
            usuarios_admpostventa = usuarios_admpostventa.filter(area='ADM POST VENTA', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_admpostventa = Usuario.objects.filter(area='ADM POST VENTA')
            usuarios_admpostventa = usuarios_admpostventa.filter(area='ADM POST VENTA', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_admpostventa = Usuario.objects.filter(area='ADM POST VENTA')
            usuarios_admpostventa = usuarios_admpostventa.filter(area='ADM POST VENTA', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_admpostventa = Usuario.objects.filter(area='ADM POST VENTA')
            usuarios_admpostventa = usuarios_admpostventa.filter(area='ADM POST VENTA', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_admpostventa = Usuario.objects.filter(area='ADM POST VENTA')
            usuarios_admpostventa = usuarios_admpostventa.filter(area='ADM POST VENTA', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])
    
    telefonos_admpostventa = []
    for usuario in usuarios_admpostventa:
        telefonos_usuario_admpostventa = usuario.telefono_set.all()
        telefonos_admpostventa.extend(telefonos_usuario_admpostventa)

    impresoras_admpostventa = []
    for usuario1 in usuarios_admpostventa:
        impresoras_usuario_admpostventa = usuario1.impresora_set.all()
        impresoras_admpostventa.extend(impresoras_usuario_admpostventa)

    notebooks_admpostventa = []
    for usuario2 in usuarios_admpostventa:
        notebooks_usuario_admpostventa = usuario2.notebook_set.all()
        notebooks_admpostventa.extend(notebooks_usuario_admpostventa)

    pcs_admpostventa = []
    for usuario3 in usuarios_admpostventa:
        pcs_usuario_admpostventa = usuario3.pc_set.all()
        pcs_admpostventa.extend(pcs_usuario_admpostventa)

    activos_admpostventa = []
    for usuario4 in usuarios_admpostventa:
        activos_usuario_admpostventa = usuario4.activoinfraestructura_set.all()
        activos_admpostventa.extend(activos_usuario_admpostventa)

    pctotal = len(pcs_admpostventa)
    teltotal = len(telefonos_admpostventa)
    imptotal = len(impresoras_admpostventa)
    notetotal = len(notebooks_admpostventa)
    acttotal = len(activos_admpostventa)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_admpostventa, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_admpostventa, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_admpostventa, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_admpostventa, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_admpostventa, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/ADMPOSTVENTA.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#ADMINISTRACION
@login_required
def administracion(request, company=None):

    usuarios_administracion = Usuario.objects.filter(area='ADMINISTRACION')
    
    usuarios_administracion = usuarios_administracion.filter(area='ADMINISTRACION', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_administracion = Usuario.objects.filter(area='ADMINISTRACION')
            usuarios_administracion = usuarios_administracion.filter(area='ADMINISTRACION', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_administracion = Usuario.objects.filter(area='ADMINISTRACION')
            usuarios_administracion = usuarios_administracion.filter(area='ADMINISTRACION', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_administracion = Usuario.objects.filter(area='ADMINISTRACION')
            usuarios_administracion = usuarios_administracion.filter(area='ADMINISTRACION', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_administracion = Usuario.objects.filter(area='ADMINISTRACION')
            usuarios_administracion = usuarios_administracion.filter(area='ADMINISTRACION', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_administracion = Usuario.objects.filter(area='ADMINISTRACION')
            usuarios_administracion = usuarios_administracion.filter(area='ADMINISTRACION', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    
    fecha = request.GET.get('fecha')
    
    if fecha:
        fecha_entrada = datetime.strptime(fecha, "%Y-%m-%d").date()
        telefonos_administracion = telefonos_administracion.filter(fecha_entrada__exact=fecha_entrada)
        impresoras_administracion = impresoras_administracion.filter(fecha_entrada__exact=fecha_entrada)
        notebooks_administracion = notebooks_administracion.filter(fecha_entrada__exact=fecha_entrada)
        pcs_administracion = pcs_administracion.filter(fecha_entrada__exact=fecha_entrada)

    telefonos_administracion = []
    for usuario in usuarios_administracion:
        telefonos_usuario_administracion = usuario.telefono_set.all()
        telefonos_administracion.extend(telefonos_usuario_administracion)

    impresoras_administracion = []
    for usuario1 in usuarios_administracion:
        impresoras_usuario_administracion = usuario1.impresora_set.all()
        impresoras_administracion.extend(impresoras_usuario_administracion)

    notebooks_administracion = []
    for usuario2 in usuarios_administracion:
        notebooks_usuario_administracion = usuario2.notebook_set.all()
        notebooks_administracion.extend(notebooks_usuario_administracion)

    pcs_administracion = []
    for usuario3 in usuarios_administracion:
        pcs_usuario_administracion = usuario3.pc_set.all()
        pcs_administracion.extend(pcs_usuario_administracion)
        
    activos_administracion = []
    for usuario4 in usuarios_administracion:
        activos_usuario_administracion = usuario4.activoinfraestructura_set.all()
        activos_administracion.extend(activos_usuario_administracion)

    pctotal = len(pcs_administracion)
    teltotal = len(telefonos_administracion)
    imptotal = len(impresoras_administracion)
    notetotal = len(notebooks_administracion)
    acttotal = len(activos_administracion)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_administracion, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_administracion, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_administracion, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_administracion, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_administracion, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/ADMINISTRACION.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})


#ADMINISTRACION CENTRAL
@login_required
def administracioncentral(request, company=None):

    usuarios_administracioncentral = Usuario.objects.filter(area='ADMINISTRACION CENTRAL')
    
    if company:
        usuarios_administracioncentral = usuarios_administracioncentral.filter(area='ADMINISTRACION CENTRAL', lab_lpg=company)


    telefonos_administracioncentral = []
    for usuario in usuarios_administracioncentral:
        telefonos_usuario_administracioncentral = usuario.telefono_set.all()
        telefonos_administracioncentral.extend(telefonos_usuario_administracioncentral)

    impresoras_administracioncentral = []
    for usuario1 in usuarios_administracioncentral:
        impresoras_usuario_administracioncentral = usuario1.impresora_set.all()
        impresoras_administracioncentral.extend(impresoras_usuario_administracioncentral)

    notebooks_administracioncentral = []
    for usuario2 in usuarios_administracioncentral:
        notebooks_usuario_administracioncentral = usuario2.notebook_set.all()
        notebooks_administracioncentral.extend(notebooks_usuario_administracioncentral)

    pcs_administracioncentral = []
    for usuario3 in usuarios_administracioncentral:
        pcs_usuario_administracioncentral = usuario3.pc_set.all()
        pcs_administracioncentral.extend(pcs_usuario_administracioncentral)


    paginator = Paginator(telefonos_administracioncentral, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_administracioncentral, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_administracioncentral, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_administracioncentral, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render(request, 'area/ADMINISTRACIONCENTRAL.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3})


#PREENT Y ALISTAJE VN
@login_required
def preentyalistajevn(request, company=None):

    usuarios_preentyalistajevn = Usuario.objects.filter(area='PREENT Y ALISTAJE VN')
    
    usuarios_preentyalistajevn = usuarios_preentyalistajevn.filter(area='PREENT Y ALISTAJE VN', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_preentyalistajevn = Usuario.objects.filter(area='PREENT Y ALISTAJE VN')
            usuarios_preentyalistajevn = usuarios_preentyalistajevn.filter(area='PREENT Y ALISTAJE VN', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_preentyalistajevn = Usuario.objects.filter(area='PREENT Y ALISTAJE VN')
            usuarios_preentyalistajevn = usuarios_preentyalistajevn.filter(area='PREENT Y ALISTAJE VN', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_preentyalistajevn = Usuario.objects.filter(area='PREENT Y ALISTAJE VN')
            usuarios_preentyalistajevn = usuarios_preentyalistajevn.filter(area='PREENT Y ALISTAJE VN', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_preentyalistajevn = Usuario.objects.filter(area='PREENT Y ALISTAJE VN')
            usuarios_preentyalistajevn = usuarios_preentyalistajevn.filter(area='PREENT Y ALISTAJE VN', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_preentyalistajevn = Usuario.objects.filter(area='PREENT Y ALISTAJE VN')
            usuarios_preentyalistajevn = usuarios_preentyalistajevn.filter(area='PREENT Y ALISTAJE VN', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_preentyalistajevn = []
    for usuario in usuarios_preentyalistajevn:
        telefonos_usuario_preentyalistajevn = usuario.telefono_set.all()
        telefonos_preentyalistajevn.extend(telefonos_usuario_preentyalistajevn)

    impresoras_preentyalistajevn = []
    for usuario1 in usuarios_preentyalistajevn:
        impresoras_usuario_preentyalistajevn = usuario1.impresora_set.all()
        impresoras_preentyalistajevn.extend(impresoras_usuario_preentyalistajevn)

    notebooks_preentyalistajevn = []
    for usuario2 in usuarios_preentyalistajevn:
        notebooks_usuario_preentyalistajevn = usuario2.notebook_set.all()
        notebooks_preentyalistajevn.extend(notebooks_usuario_preentyalistajevn)

    pcs_preentyalistajevn = []
    for usuario3 in usuarios_preentyalistajevn:
        pcs_usuario_preentyalistajevn = usuario3.pc_set.all()
        pcs_preentyalistajevn.extend(pcs_usuario_preentyalistajevn)
        
    activos_preentyalistajevn = []
    for usuario4 in usuarios_preentyalistajevn:
        activos_usuario_preentyalistajevn = usuario4.activoinfraestructura_set.all()
        activos_preentyalistajevn.extend(activos_usuario_preentyalistajevn)
        
    pctotal = len(pcs_preentyalistajevn)
    teltotal = len(telefonos_preentyalistajevn)
    imptotal = len(impresoras_preentyalistajevn)
    notetotal = len(notebooks_preentyalistajevn)
    acttotal = len(activos_preentyalistajevn)

    total = pctotal + teltotal + imptotal + notetotal + acttotal


    paginator = Paginator(telefonos_preentyalistajevn, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_preentyalistajevn, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_preentyalistajevn, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_preentyalistajevn, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_preentyalistajevn, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/PREENTYALISTAJEVN.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})


#POST VENTA SERVICIOS
@login_required
def postventaservicios(request, company=None):

    usuarios_postventaservicios = Usuario.objects.filter(area='POST VENTA SERVICIOS')
    
    usuarios_postventaservicios = usuarios_postventaservicios.filter(area='POST VENTA SERVICIOS', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_postventaservicios = Usuario.objects.filter(area='POST VENTA SERVICIOS')
            usuarios_postventaservicios = usuarios_postventaservicios.filter(area='POST VENTA SERVICIOS', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_postventaservicios = Usuario.objects.filter(area='POST VENTA SERVICIOS')
            usuarios_postventaservicios = usuarios_postventaservicios.filter(area='POST VENTA SERVICIOS', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_postventaservicios = Usuario.objects.filter(area='POST VENTA SERVICIOS')
            usuarios_postventaservicios = usuarios_postventaservicios.filter(area='POST VENTA SERVICIOS', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_postventaservicios = Usuario.objects.filter(area='POST VENTA SERVICIOS')
            usuarios_postventaservicios = usuarios_postventaservicios.filter(area='POST VENTA SERVICIOS', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_postventaservicios = Usuario.objects.filter(area='POST VENTA SERVICIOS')
            usuarios_postventaservicios = usuarios_postventaservicios.filter(area='POST VENTA SERVICIOS', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    telefonos_postventaservicios = []
    for usuario in usuarios_postventaservicios:
        telefonos_usuario_postventaservicios = usuario.telefono_set.all()
        telefonos_postventaservicios.extend(telefonos_usuario_postventaservicios)

    impresoras_postventaservicios = []
    for usuario1 in usuarios_postventaservicios:
        impresoras_usuario_postventaservicios = usuario1.impresora_set.all()
        impresoras_postventaservicios.extend(impresoras_usuario_postventaservicios)

    notebooks_postventaservicios = []
    for usuario2 in usuarios_postventaservicios:
        notebooks_usuario_postventaservicios = usuario2.notebook_set.all()
        notebooks_postventaservicios.extend(notebooks_usuario_postventaservicios)

    pcs_postventaservicios = []
    for usuario3 in usuarios_postventaservicios:
        pcs_usuario_postventaservicios = usuario3.pc_set.all()
        pcs_postventaservicios.extend(pcs_usuario_postventaservicios)
        
    activos_postventaservicios = []
    for usuario4 in usuarios_postventaservicios:
        activos_usuario_admpostventaservicios = usuario4.activoinfraestructura_set.all()
        activos_postventaservicios.extend(activos_usuario_admpostventaservicios)
        
    pctotal = len(pcs_postventaservicios)
    teltotal = len(telefonos_postventaservicios)
    imptotal = len(impresoras_postventaservicios)
    notetotal = len(notebooks_postventaservicios)
    acttotal = len(activos_postventaservicios)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_postventaservicios, 800)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_postventaservicios, 800)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_postventaservicios, 800)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_postventaservicios, 800)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_postventaservicios, 800)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render (request, 'area/POSTVENTASERVICIOS.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#REPUESTOS SUCURSAL
@login_required
def repuestossucursal(request, company=None):
    usuarios_repuestossucursal = Usuario.objects.filter(area='REPUESTOS SUCURSAL')
    
    if company:
        usuarios_repuestossucursal = usuarios_repuestossucursal.filter(area='REPUESTOS SUCURSAL', lab_lpg=company)

    telefonos_repuestossucursal = []
    for usuario in usuarios_repuestossucursal:
        telefonos_usuario_repuestossucursal = usuario.telefono_set.all()
        telefonos_repuestossucursal.extend(telefonos_usuario_repuestossucursal)

    impresoras_repuestossucursal = []
    for usuario1 in usuarios_repuestossucursal:
        impresoras_usuario_repuestossucursal = usuario1.impresora_set.all()
        impresoras_repuestossucursal.extend(impresoras_usuario_repuestossucursal)

    notebooks_repuestossucursal = []
    for usuario2 in usuarios_repuestossucursal:
        notebooks_usuario_repuestossucursal = usuario2.notebook_set.all()
        notebooks_repuestossucursal.extend(notebooks_usuario_repuestossucursal)

    pcs_repuestossucursal = []
    for usuario3 in usuarios_repuestossucursal:
        pcs_usuario_repuestossucursal = usuario3.pc_set.all()
        pcs_repuestossucursal.extend(pcs_usuario_repuestossucursal)

    paginator = Paginator(telefonos_repuestossucursal, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_repuestossucursal, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_repuestossucursal, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_repuestossucursal, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render (request, 'area/REPUESTOSSUCURSAL.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3})

#TALLER SUCURSAL
@login_required
def tallersucursal(request, company=None):

    usuarios_tallersucursal = Usuario.objects.filter(area='TALLER SUCURSAL')
    
    if company:
        usuarios_tallersucursal = usuarios_tallersucursal.filter(area='TALLER SUCURSAL', lab_lpg=company)


    telefonos_tallersucursal = []
    for usuario in usuarios_tallersucursal:
        telefonos_usuario_tallersucursal = usuario.telefono_set.all()
        telefonos_tallersucursal.extend(telefonos_usuario_tallersucursal)

    impresoras_tallersucursal = []
    for usuario1 in usuarios_tallersucursal:
        impresoras_usuario_tallersucursal = usuario1.impresora_set.all()
        impresoras_tallersucursal.extend(impresoras_usuario_tallersucursal)

    notebooks_tallersucursal = []
    for usuario2 in usuarios_tallersucursal:
        notebooks_usuario_tallersucursal = usuario2.notebook_set.all()
        notebooks_tallersucursal.extend(notebooks_usuario_tallersucursal)

    pcs_tallersucursal = []
    for usuario3 in usuarios_tallersucursal:
        pcs_usuario_tallersucursal = usuario3.pc_set.all()
        pcs_tallersucursal.extend(pcs_usuario_tallersucursal)


    paginator = Paginator(telefonos_tallersucursal, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_tallersucursal, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_tallersucursal, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_tallersucursal, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render(request, 'area/TALLERSUCURSAL.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3})

#PLANDEAHORRO
@login_required
def plandeahorro(request, company=None):

    usuarios_plandeahorro = Usuario.objects.filter(area='PLAN DE AHORRO')
    
    usuarios_plandeahorro = usuarios_plandeahorro.filter(area='PLAN DE AHORRO', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_plandeahorro = Usuario.objects.filter(area='PLAN DE AHORRO')
            usuarios_plandeahorro = usuarios_plandeahorro.filter(area='PLAN DE AHORRO', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_plandeahorro = Usuario.objects.filter(area='PLAN DE AHORRO')
            usuarios_plandeahorro = usuarios_plandeahorro.filter(area='PLAN DE AHORRO', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_plandeahorro = Usuario.objects.filter(area='PLAN DE AHORRO')
            usuarios_plandeahorro = usuarios_plandeahorro.filter(area='PLAN DE AHORRO', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_plandeahorro = Usuario.objects.filter(area='PLAN DE AHORRO')
            usuarios_plandeahorro = usuarios_plandeahorro.filter(area='PLAN DE AHORRO', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_plandeahorro = Usuario.objects.filter(area='PLAN DE AHORRO')
            usuarios_plandeahorro = usuarios_plandeahorro.filter(area='PLAN DE AHORRO', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])
    
    

    totalusuarios = usuarios_plandeahorro.count()

    telefonos_plandeahorro = []
    for usuario in usuarios_plandeahorro:
        telefonos_usuario_plandeahorro = usuario.telefono_set.all()
        telefonos_plandeahorro.extend(telefonos_usuario_plandeahorro)

    impresoras_plandeahorro = []
    for usuario1 in usuarios_plandeahorro:
        impresoras_usuario_plandeahorro = usuario1.impresora_set.all()
        impresoras_plandeahorro.extend(impresoras_usuario_plandeahorro)

    notebooks_plandeahorro = []
    for usuario2 in usuarios_plandeahorro:
        notebooks_usuario_plandeahorro = usuario2.notebook_set.all()
        notebooks_plandeahorro.extend(notebooks_usuario_plandeahorro)

    pcs_plandeahorro = []
    for usuario3 in usuarios_plandeahorro:
        pcs_usuario_plandeahorro = usuario3.pc_set.all()
        pcs_plandeahorro.extend(pcs_usuario_plandeahorro)
        
    activos_plandeahorro = []
    for usuario4 in usuarios_plandeahorro:
        activos_usuario_plandeahorro = usuario4.activoinfraestructura_set.all()
        activos_plandeahorro.extend(activos_usuario_plandeahorro)
        
    pctotal = len(pcs_plandeahorro)
    teltotal = len(telefonos_plandeahorro)
    imptotal = len(impresoras_plandeahorro)
    notetotal = len(notebooks_plandeahorro)
    acttotal = len(activos_plandeahorro)

    total = pctotal + teltotal + imptotal + notetotal + acttotal


    paginator = Paginator(telefonos_plandeahorro, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_plandeahorro, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_plandeahorro, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_plandeahorro, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(pcs_plandeahorro, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/PLANDEAHORRO.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'totalusuarios':totalusuarios,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})


#PLAN DE AHORRO SUCURSAL
@login_required
def plandeahorrosucursal(request, company=None):

    usuarios_plandeahorrosucursal = Usuario.objects.filter(area='PLAN DE AHORRO SUCURSAL')
    
    if company:
        usuarios_plandeahorrosucursal = usuarios_plandeahorrosucursal.filter(area='PLAN DE AHORRO SUCURSAL', lab_lpg=company)

    telefonos_plandeahorrosucursal = []
    for usuario in usuarios_plandeahorrosucursal:
        telefonos_usuario_plandeahorrosucursal = usuario.telefono_set.all()
        telefonos_plandeahorrosucursal.extend(telefonos_usuario_plandeahorrosucursal)

    impresoras_plandeahorrosucursal = []
    for usuario1 in usuarios_plandeahorrosucursal:
        impresoras_usuario_plandeahorrosucursal = usuario1.impresora_set.all()
        impresoras_plandeahorrosucursal.extend(impresoras_usuario_plandeahorrosucursal)

    notebooks_plandeahorrosucursal = []
    for usuario2 in usuarios_plandeahorrosucursal:
        notebooks_usuario_plandeahorrosucursal = usuario2.notebook_set.all()
        notebooks_plandeahorrosucursal.extend(notebooks_usuario_plandeahorrosucursal)

    pcs_plandeahorrosucursal = []
    for usuario3 in usuarios_plandeahorrosucursal:
        pcs_usuario_plandeahorrosucursal = usuario3.pc_set.all()
        pcs_plandeahorrosucursal.extend(pcs_usuario_plandeahorrosucursal)

    paginator = Paginator(telefonos_plandeahorrosucursal, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_plandeahorrosucursal, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_plandeahorrosucursal, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_plandeahorrosucursal, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render (request, 'area/PLANDEAHORROSUCURSAL.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3})

#SISTEMAS
@login_required
def sistemas(request, company=None):

    usuarios_sistema_helpdesk = Usuario.objects.filter(area='SISTEMAS')
    
    usuarios_sistema_helpdesk = usuarios_sistema_helpdesk.filter(area='SISTEMAS', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_sistema_helpdesk = Usuario.objects.filter(area='SISTEMAS')
            usuarios_sistema_helpdesk = usuarios_sistema_helpdesk.filter(area='SISTEMAS', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_sistema_helpdesk = Usuario.objects.filter(area='SISTEMAS')
            usuarios_sistema_helpdesk = usuarios_sistema_helpdesk.filter(area='SISTEMAS', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_sistema_helpdesk = Usuario.objects.filter(area='SISTEMAS')
            usuarios_sistema_helpdesk = usuarios_sistema_helpdesk.filter(area='SISTEMAS', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_sistema_helpdesk = Usuario.objects.filter(area='SISTEMAS')
            usuarios_sistema_helpdesk = usuarios_sistema_helpdesk.filter(area='SISTEMAS', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_sistema_helpdesk = Usuario.objects.filter(area='SISTEMAS')
            usuarios_sistema_helpdesk = usuarios_sistema_helpdesk.filter(area='SISTEMAS', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    telefonos_sistemas = []
    for usuario in usuarios_sistema_helpdesk:
        telefonos_usuario_sistemas = usuario.telefono_set.all()
        telefonos_sistemas.extend(telefonos_usuario_sistemas)

    impresoras_sistemas = []
    for usuario1 in usuarios_sistema_helpdesk:
        impresoras_usuario_sistemas = usuario1.impresora_set.all()
        impresoras_sistemas.extend(impresoras_usuario_sistemas)

    notebooks_sistemas = []
    for usuario2 in usuarios_sistema_helpdesk:
        notebooks_usuario_sistemas = usuario2.notebook_set.all()
        notebooks_sistemas.extend(notebooks_usuario_sistemas)

    pcs_sistemas = []
    for usuario3 in usuarios_sistema_helpdesk:
        pcs_usuario_sistemas = usuario3.pc_set.all()
        pcs_sistemas.extend(pcs_usuario_sistemas)

    activos_sistemas = []
    for usuario4 in usuarios_sistema_helpdesk:
        activos_usuario_sistemas = usuario4.activoinfraestructura_set.all()
        activos_sistemas.extend(activos_usuario_sistemas)

    pctotal = len(pcs_sistemas)
    teltotal = len(telefonos_sistemas)
    imptotal = len(impresoras_sistemas)
    notetotal = len(notebooks_sistemas)
    acttotal = len(activos_sistemas)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_sistemas, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_sistemas, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_sistemas, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_sistemas, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_sistemas, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render (request, 'area/SISTEMAS.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#LAVADERO
@login_required
def lavadero(request, company=None):

    usuarios_lavadero = Usuario.objects.filter(area='LAVADERO')

    if company:
        usuarios_lavadero = usuarios_lavadero.filter(area='LAVADERO', lab_lpg=company)


    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)


    paginator = Paginator(telefonos_lavadero, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)


    return render (request, 'area/LAVADERO.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3})

#TALLER
@login_required
def taller(request, company=None):

    usuarios_taller = Usuario.objects.filter(area='TALLER')
    
    usuarios_taller = usuarios_taller.filter(area='TALLER', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_taller = Usuario.objects.filter(area='TALLER')
            usuarios_taller = usuarios_taller.filter(area='TALLER', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_taller = Usuario.objects.filter(area='TALLER')
            usuarios_taller = usuarios_taller.filter(area='TALLER', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_taller = Usuario.objects.filter(area='TALLER')
            usuarios_taller = usuarios_taller.filter(area='TALLER', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_taller = Usuario.objects.filter(area='TALLER')
            usuarios_taller = usuarios_taller.filter(area='TALLER', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_taller = Usuario.objects.filter(area='TALLER')
            usuarios_taller = usuarios_taller.filter(area='TALLER', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_taller = []
    for usuario in usuarios_taller:
        telefonos_usuario_taller = usuario.telefono_set.all()
        telefonos_taller.extend(telefonos_usuario_taller)

    impresoras_taller = []
    for usuario1 in usuarios_taller:
        impresoras_usuario_taller = usuario1.impresora_set.all()
        impresoras_taller.extend(impresoras_usuario_taller)

    notebooks_taller = []
    for usuario2 in usuarios_taller:
        notebooks_usuario_taller = usuario2.notebook_set.all()
        notebooks_taller.extend(notebooks_usuario_taller)

    pcs_taller = []
    for usuario3 in usuarios_taller:
        pcs_usuario_taller = usuario3.pc_set.all()
        pcs_taller.extend(pcs_usuario_taller)

    activos_taller = []
    for usuario4 in usuarios_taller:
        activos_usuario_taller = usuario4.activoinfraestructura_set.all()
        activos_taller.extend(activos_usuario_taller)

    pctotal = len(pcs_taller)
    teltotal = len(telefonos_taller)
    imptotal = len(impresoras_taller)
    notetotal = len(notebooks_taller)
    acttotal = len(activos_taller)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_taller, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_taller, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_taller, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_taller, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_taller, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/TALLER.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#MKT
@login_required
def mkt(request, company=None):

    usuarios_mkt = Usuario.objects.filter(area='MKT')
    
    usuarios_mkt = usuarios_mkt.filter(area='MKT', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_mkt = Usuario.objects.filter(area='MKT')
            usuarios_mkt = usuarios_mkt.filter(area='MKT', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_mkt = Usuario.objects.filter(area='MKT')
            usuarios_mkt = usuarios_mkt.filter(area='MKT', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_mkt = Usuario.objects.filter(area='MKT')
            usuarios_mkt = usuarios_mkt.filter(area='MKT', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_mkt = Usuario.objects.filter(area='MKT')
            usuarios_mkt = usuarios_mkt.filter(area='MKT', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_mkt = Usuario.objects.filter(area='MKT')
            usuarios_mkt = usuarios_mkt.filter(area='MKT', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_mkt = []
    for usuario in usuarios_mkt:
        telefonos_usuario_mkt = usuario.telefono_set.all()
        telefonos_mkt.extend(telefonos_usuario_mkt)

    impresoras_mkt = []
    for usuario1 in usuarios_mkt:
        impresoras_usuario_mkt = usuario1.impresora_set.all()
        impresoras_mkt.extend(impresoras_usuario_mkt)

    notebooks_mkt = []
    for usuario2 in usuarios_mkt:
        notebooks_usuario_mkt = usuario2.notebook_set.all()
        notebooks_mkt.extend(notebooks_usuario_mkt)

    pcs_mkt = []
    for usuario3 in usuarios_mkt:
        pcs_usuario_mkt = usuario3.pc_set.all()
        pcs_mkt.extend(pcs_usuario_mkt)
        
    activos_mkt = []
    for usuario4 in usuarios_mkt:
        activos_usuario_mkt = usuario4.activoinfraestructura_set.all()
        activos_mkt.extend(activos_usuario_mkt)
        
    pctotal = len(pcs_mkt)
    teltotal = len(telefonos_mkt)
    imptotal = len(impresoras_mkt)
    notetotal = len(notebooks_mkt)
    acttotal = len(activos_mkt)

    total = pctotal + teltotal + imptotal + notetotal + acttotal
        
    paginator = Paginator(telefonos_mkt, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_mkt, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_mkt, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_mkt, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_mkt, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/MKT.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'acttotal':acttotal,'notetotal':notetotal,'imptotal':imptotal,'teltotal':teltotal,'pctotal':pctotal,'total':total})

#COMERCIAL
@login_required
def comercial(request, company=None):

    usuarios_comercial = Usuario.objects.filter(area='COMERCIAL')
    
    usuarios_comercial = usuarios_comercial.filter(area='COMERCIAL', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_comercial = Usuario.objects.filter(area='COMERCIAL')
            usuarios_comercial = usuarios_comercial.filter(area='COMERCIAL', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_comercial = Usuario.objects.filter(area='COMERCIAL')
            usuarios_comercial = usuarios_comercial.filter(area='COMERCIAL', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_comercial = Usuario.objects.filter(area='COMERCIAL')
            usuarios_comercial = usuarios_comercial.filter(area='COMERCIAL', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_comercial = Usuario.objects.filter(area='COMERCIAL')
            usuarios_comercial = usuarios_comercial.filter(area='COMERCIAL', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_comercial = Usuario.objects.filter(area='COMERCIAL')
            usuarios_comercial = usuarios_comercial.filter(area='COMERCIAL', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_comercial = []
    for usuario in usuarios_comercial:
        telefonos_usuario_comercial = usuario.telefono_set.all()
        telefonos_comercial.extend(telefonos_usuario_comercial)

    impresoras_comercial = []
    for usuario1 in usuarios_comercial:
        impresoras_usuario_comercial = usuario1.impresora_set.all()
        impresoras_comercial.extend(impresoras_usuario_comercial)

    notebooks_comercial = []
    for usuario2 in usuarios_comercial:
        notebooks_usuario_comercial = usuario2.notebook_set.all()
        notebooks_comercial.extend(notebooks_usuario_comercial)

    pcs_comercial = []
    for usuario3 in usuarios_comercial:
        pcs_usuario_comercial = usuario3.pc_set.all()
        pcs_comercial.extend(pcs_usuario_comercial)
        
    activos_comercial = []
    for usuario4 in usuarios_comercial:
        activos_usuario_comercial = usuario4.activoinfraestructura_set.all()
        activos_comercial.extend(activos_usuario_comercial)
        
        
    pctotal = len(pcs_comercial)
    teltotal = len(telefonos_comercial)
    imptotal = len(impresoras_comercial)
    notetotal = len(notebooks_comercial)
    acttotal = len(activos_comercial)

    total = pctotal + teltotal + imptotal + notetotal + acttotal


    paginator = Paginator(telefonos_comercial, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_comercial, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_comercial, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_comercial, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_comercial, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/COMERCIAL.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#RRHH
@login_required
def rrhh(request, company=None):
    
    usuarios_admventasplanes = Usuario.objects.filter(area='RRHH')
    
    usuarios_admventasplanes = usuarios_admventasplanes.filter(area='RRHH', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_admventasplanes = Usuario.objects.filter(area='RRHH')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='RRHH', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_admventasplanes = Usuario.objects.filter(area='RRHH')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='RRHH', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_admventasplanes = Usuario.objects.filter(area='RRHH')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='RRHH', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_admventasplanes = Usuario.objects.filter(area='RRHH')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='RRHH', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_admventasplanes = Usuario.objects.filter(area='RRHH')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='RRHH', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])
    
    
    telefonos_admventasplanes = []
    for usuario in usuarios_admventasplanes:
        telefonos_usuario_admventasplanes = usuario.telefono_set.all()
        telefonos_admventasplanes.extend(telefonos_usuario_admventasplanes)

    impresoras_admventasplanes = []
    for usuario1 in usuarios_admventasplanes:
        impresoras_usuario_admventasplanes = usuario1.impresora_set.all()
        impresoras_admventasplanes.extend(impresoras_usuario_admventasplanes)

    notebooks_admventasplanes = []
    for usuario2 in usuarios_admventasplanes:
        notebooks_usuario_admventasplanes = usuario2.notebook_set.all()
        notebooks_admventasplanes.extend(notebooks_usuario_admventasplanes)

    pcs_admventasplanes = []
    for usuario3 in usuarios_admventasplanes:
        pcs_usuario_admventasplanes = usuario3.pc_set.all()
        pcs_admventasplanes.extend(pcs_usuario_admventasplanes)
        
    activos_admventasplanes = []
    for usuario4 in usuarios_admventasplanes:
        activos_usuario_admventasplanes = usuario4.activoinfraestructura_set.all()
        activos_admventasplanes.extend(activos_usuario_admventasplanes)
        
        
    pctotal = len(pcs_admventasplanes)
    teltotal = len(telefonos_admventasplanes)
    imptotal = len(impresoras_admventasplanes)
    notetotal = len(notebooks_admventasplanes)
    acttotal = len(activos_admventasplanes)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/RRHH.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'notetotal':notetotal,'imptotal':imptotal,'teltotal':teltotal,'acttotal':acttotal,'total':total})

#GESTORIA
@login_required
def gestoria(request, company=None):

    usuarios_gestoria = Usuario.objects.filter(area='GESTORIA')
    
    if company:
        usuarios_gestoria = usuarios_gestoria.filter(area='GESTORIA', lab_lpg=company)


    telefonos_gestoria = []
    for usuario in usuarios_gestoria:
        telefonos_usuario_gestoria = usuario.telefono_set.all()
        telefonos_gestoria.extend(telefonos_usuario_gestoria)

    impresoras_gestoria = []
    for usuario1 in usuarios_gestoria:
        impresoras_usuario_gestoria = usuario1.impresora_set.all()
        impresoras_gestoria.extend(impresoras_usuario_gestoria)

    notebooks_gestoria = []
    for usuario2 in usuarios_gestoria:
        notebooks_usuario_gestoria = usuario2.notebook_set.all()
        notebooks_gestoria.extend(notebooks_usuario_gestoria)

    pcs_gestoria = []
    for usuario3 in usuarios_gestoria:
        pcs_usuario_gestoria = usuario3.pc_set.all()
        pcs_gestoria.extend(pcs_usuario_gestoria)


    paginator = Paginator(telefonos_gestoria, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_gestoria, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_gestoria, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_gestoria, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render(request, 'area/GESTORIA.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3})

#MAESTRANZA
@login_required
def maestranza(request, company=None):

    usuarios_maestranza = Usuario.objects.filter(area='MAESTRANZA')
    
    if company:
        usuarios_maestranza = usuarios_maestranza.filter(area='MAESTRANZA', lab_lpg=company)


    telefonos_maestranza = []
    for usuario in usuarios_maestranza:
        telefonos_usuario_maestranza = usuario.telefono_set.all()
        telefonos_maestranza.extend(telefonos_usuario_maestranza)

    impresoras_maestranza = []
    for usuario1 in usuarios_maestranza:
        impresoras_usuario_maestranza = usuario1.impresora_set.all()
        impresoras_maestranza.extend(impresoras_usuario_maestranza)

    notebooks_maestranza = []
    for usuario2 in usuarios_maestranza:
        notebooks_usuario_maestranza = usuario2.notebook_set.all()
        notebooks_maestranza.extend(notebooks_usuario_maestranza)

    pcs_maestranza = []
    for usuario3 in usuarios_maestranza:
        pcs_usuario_maestranza = usuario3.pc_set.all()
        pcs_maestranza.extend(pcs_usuario_maestranza)


    paginator = Paginator(telefonos_maestranza, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_maestranza, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_maestranza, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_maestranza, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render(request, 'area/MAESTRANZA.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3})

#VENTA DIRECTA VO
@login_required
def ventadirectavo(request, company=None):

    usuarios_ventadirectavo = Usuario.objects.filter(area='VENTA DIRECTA VO')
    
    usuarios_ventadirectavo = usuarios_ventadirectavo.filter(area='VENTA DIRECTA VO', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_ventadirectavo = Usuario.objects.filter(area='VENTA DIRECTA VO')
            usuarios_ventadirectavo = usuarios_ventadirectavo.filter(area='VENTA DIRECTA VO', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_ventadirectavo = Usuario.objects.filter(area='VENTA DIRECTA VO')
            usuarios_ventadirectavo = usuarios_ventadirectavo.filter(area='VENTA DIRECTA VO', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_ventadirectavo = Usuario.objects.filter(area='VENTA DIRECTA VO')
            usuarios_ventadirectavo = usuarios_ventadirectavo.filter(area='VENTA DIRECTA VO', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_ventadirectavo = Usuario.objects.filter(area='VENTA DIRECTA VO')
            usuarios_ventadirectavo = usuarios_ventadirectavo.filter(area='VENTA DIRECTA VO', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_ventadirectavo = Usuario.objects.filter(area='VENTA DIRECTA VO')
            usuarios_ventadirectavo = usuarios_ventadirectavo.filter(area='VENTA DIRECTA VO', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    telefonos_ventadirectavo = []
    for usuario in usuarios_ventadirectavo:
        telefonos_usuario_ventadirectavo = usuario.telefono_set.all()
        telefonos_ventadirectavo.extend(telefonos_usuario_ventadirectavo)

    impresoras_ventadirectavo = []
    for usuario1 in usuarios_ventadirectavo:
        impresoras_usuario_ventadirectavo = usuario1.impresora_set.all()
        impresoras_ventadirectavo.extend(impresoras_usuario_ventadirectavo)

    notebooks_ventadirectavo = []
    for usuario2 in usuarios_ventadirectavo:
        notebooks_usuario_ventadirectavo = usuario2.notebook_set.all()
        notebooks_ventadirectavo.extend(notebooks_usuario_ventadirectavo)

    pcs_ventadirectavo = []
    for usuario3 in usuarios_ventadirectavo:
        pcs_usuario_ventadirectavo = usuario3.pc_set.all()
        pcs_ventadirectavo.extend(pcs_usuario_ventadirectavo)

    activos_ventadirectavo = []
    for usuario4 in usuarios_ventadirectavo:
        activos_usuario_ventadirectavo = usuario4.activoinfraestructura_set.all()
        activos_ventadirectavo.extend(activos_usuario_ventadirectavo)

    pctotal = len(pcs_ventadirectavo)
    teltotal = len(telefonos_ventadirectavo)
    imptotal = len(impresoras_ventadirectavo)
    notetotal = len(notebooks_ventadirectavo)
    acttotal = len(activos_ventadirectavo)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_ventadirectavo, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_ventadirectavo, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_ventadirectavo, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_ventadirectavo, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_ventadirectavo, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/VENTADIRECTAVO.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#PREENT Y ALISTAJE VO
@login_required
def preentyalistajevo(request, company=None):

    usuarios_preentyalistajevo = Usuario.objects.filter(area='PREENT Y ALISTAJE VO')
    
    usuarios_preentyalistajevo = usuarios_preentyalistajevo.filter(area='PREENT Y ALISTAJE VO', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_preentyalistajevo = Usuario.objects.filter(area='APREENT Y ALISTAJE VO')
            usuarios_preentyalistajevo = usuarios_preentyalistajevo.filter(area='APREENT Y ALISTAJE VO', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_preentyalistajevo = Usuario.objects.filter(area='APREENT Y ALISTAJE VO')
            usuarios_preentyalistajevo = usuarios_preentyalistajevo.filter(area='APREENT Y ALISTAJE VO', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_preentyalistajevo = Usuario.objects.filter(area='APREENT Y ALISTAJE VO')
            usuarios_preentyalistajevo = usuarios_preentyalistajevo.filter(area='APREENT Y ALISTAJE VO', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_preentyalistajevo = Usuario.objects.filter(area='APREENT Y ALISTAJE VO')
            usuarios_preentyalistajevo = usuarios_preentyalistajevo.filter(area='APREENT Y ALISTAJE VO', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_preentyalistajevo = Usuario.objects.filter(area='APREENT Y ALISTAJE VO')
            usuarios_preentyalistajevo = usuarios_preentyalistajevo.filter(area='APREENT Y ALISTAJE VO', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_preentyalistajevo = []
    for usuario in usuarios_preentyalistajevo:
        telefonos_usuario_preentyalistajevo = usuario.telefono_set.all()
        telefonos_preentyalistajevo.extend(telefonos_usuario_preentyalistajevo)

    impresoras_preentyalistajevo = []
    for usuario1 in usuarios_preentyalistajevo:
        impresoras_usuario_preentyalistajevo = usuario1.impresora_set.all()
        impresoras_preentyalistajevo.extend(impresoras_usuario_preentyalistajevo)

    notebooks_preentyalistajevo = []
    for usuario2 in usuarios_preentyalistajevo:
        notebooks_usuario_preentyalistajevo = usuario2.notebook_set.all()
        notebooks_preentyalistajevo.extend(notebooks_usuario_preentyalistajevo)

    pcs_preentyalistajevo = []
    for usuario3 in usuarios_preentyalistajevo:
        pcs_usuario_preentyalistajevo = usuario3.pc_set.all()
        pcs_preentyalistajevo.extend(pcs_usuario_preentyalistajevo)

    activos_preentyalistajevo = []
    for usuario4 in usuarios_preentyalistajevo:
        activos_usuario_preentyalistajevo = usuario4.activoinfraestructura_set.all()
        activos_preentyalistajevo.extend(activos_usuario_preentyalistajevo)

    pctotal = len(pcs_preentyalistajevo)
    teltotal = len(telefonos_preentyalistajevo)
    imptotal = len(impresoras_preentyalistajevo)
    notetotal = len(notebooks_preentyalistajevo)
    acttotal = len(activos_preentyalistajevo)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_preentyalistajevo, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_preentyalistajevo, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_preentyalistajevo, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_preentyalistajevo, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_preentyalistajevo, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/PREENTYALISTAJEVO.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#ADMINISTRACION SUCURSAL
@login_required
def administracionsucursal(request, company=None):

    usuarios_administracionsucursal = Usuario.objects.filter(area='ADMINISTRACION SUCURSAL')
    
    if company:
        usuarios_administracionsucursal = usuarios_administracionsucursal.filter(area='ADMINISTRACION SUCURSAL', lab_lpg=company)


    telefonos_administracionsucursal = []
    for usuario in usuarios_administracionsucursal:
        telefonos_usuario_administracionsucursal = usuario.telefono_set.all()
        telefonos_administracionsucursal.extend(telefonos_usuario_administracionsucursal)

    impresoras_administracionsucursal = []
    for usuario1 in usuarios_administracionsucursal:
        impresoras_usuario_administracionsucursal = usuario1.impresora_set.all()
        impresoras_administracionsucursal.extend(impresoras_usuario_administracionsucursal)

    notebooks_administracionsucursal = []
    for usuario2 in usuarios_administracionsucursal:
        notebooks_usuario_administracionsucursal = usuario2.notebook_set.all()
        notebooks_administracionsucursal.extend(notebooks_usuario_administracionsucursal)

    pcs_administracionsucursal = []
    for usuario3 in usuarios_administracionsucursal:
        pcs_usuario_administracionsucursal = usuario3.pc_set.all()
        pcs_administracionsucursal.extend(pcs_usuario_administracionsucursal)


    paginator = Paginator(telefonos_administracionsucursal, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_administracionsucursal, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_administracionsucursal, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_administracionsucursal, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render(request, 'area/ADMINISTRACIONSUCURSAL.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3})

#ADM POST VENTA SUCURSAL
@login_required
def admpostventasucursal(request, company=None):

    usuarios_admpostventasucursal = Usuario.objects.filter(area='ADM POST VENTA SUCURSAL')
    
    if company:
        usuarios_admpostventasucursal = usuarios_admpostventasucursal.filter(area='ADM POST VENTA SUCURSAL', lab_lpg=company)


    telefonos_admpostventasucursal = []
    for usuario in usuarios_admpostventasucursal:
        telefonos_usuario_admpostventasucursal = usuario.telefono_set.all()
        telefonos_admpostventasucursal.extend(telefonos_usuario_admpostventasucursal)

    impresoras_admpostventasucursal = []
    for usuario1 in usuarios_admpostventasucursal:
        impresoras_usuario_admpostventasucursal = usuario1.impresora_set.all()
        impresoras_admpostventasucursal.extend(impresoras_usuario_admpostventasucursal)

    notebooks_admpostventasucursal = []
    for usuario2 in usuarios_admpostventasucursal:
        notebooks_usuario_admpostventasucursal = usuario2.notebook_set.all()
        notebooks_admpostventasucursal.extend(notebooks_usuario_admpostventasucursal)

    pcs_admpostventasucursal = []
    for usuario3 in usuarios_admpostventasucursal:
        pcs_usuario_admpostventasucursal = usuario3.pc_set.all()
        pcs_admpostventasucursal.extend(pcs_usuario_admpostventasucursal)


    paginator = Paginator(telefonos_admpostventasucursal, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_admpostventasucursal, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_admpostventasucursal, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_admpostventasucursal, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    return render(request, 'area/ADMPOSTVENTASUCURSAL.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3})

#GERENTE COMERCIAL
@login_required
def gerentecomercial(request, company=None):

    usuarios_gerentecomercial = Usuario.objects.filter(area='GERENTE COMERCIAL')
    
    usuarios_gerentecomercial = usuarios_gerentecomercial.filter(area='GERENTE COMERCIAL', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_gerentecomercial = Usuario.objects.filter(area='GERENTE COMERCIAL')
            usuarios_gerentecomercial = usuarios_gerentecomercial.filter(area='GERENTE COMERCIAL', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_gerentecomercial = Usuario.objects.filter(area='GERENTE COMERCIAL')
            usuarios_gerentecomercial = usuarios_gerentecomercial.filter(area='GERENTE COMERCIAL', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_gerentecomercial = Usuario.objects.filter(area='GERENTE COMERCIAL')
            usuarios_gerentecomercial = usuarios_gerentecomercial.filter(area='GERENTE COMERCIAL', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_gerentecomercial = Usuario.objects.filter(area='GERENTE COMERCIAL')
            usuarios_gerentecomercial = usuarios_gerentecomercial.filter(area='GERENTE COMERCIAL', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_gerentecomercial = Usuario.objects.filter(area='GERENTE COMERCIAL')
            usuarios_gerentecomercial = usuarios_gerentecomercial.filter(area='GERENTE COMERCIAL', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_gerentecomercial = []
    for usuario in usuarios_gerentecomercial:
        telefonos_usuario_gerentecomercial = usuario.telefono_set.all()
        telefonos_gerentecomercial.extend(telefonos_usuario_gerentecomercial)

    impresoras_gerentecomercial = []
    for usuario1 in usuarios_gerentecomercial:
        impresoras_usuario_gerentecomercial = usuario1.impresora_set.all()
        impresoras_gerentecomercial.extend(impresoras_usuario_gerentecomercial)

    notebooks_gerentecomercial = []
    for usuario2 in usuarios_gerentecomercial:
        notebooks_usuario_gerentecomercial = usuario2.notebook_set.all()
        notebooks_gerentecomercial.extend(notebooks_usuario_gerentecomercial)

    pcs_gerentecomercial = []
    for usuario3 in usuarios_gerentecomercial:
        pcs_usuario_gerentecomercial = usuario3.pc_set.all()
        pcs_gerentecomercial.extend(pcs_usuario_gerentecomercial)
        
    activos_gerentecomercial = []
    for usuario4 in usuarios_gerentecomercial:
        activos_usuario_gerentecomercial = usuario4.activoinfraestructura_set.all()
        activos_gerentecomercial.extend(activos_gerentecomercial)
        
    pctotal = len(pcs_gerentecomercial)
    teltotal = len(telefonos_gerentecomercial)
    imptotal = len(impresoras_gerentecomercial)
    notetotal = len(notebooks_gerentecomercial)
    acttotal = len(activos_gerentecomercial)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_gerentecomercial, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_gerentecomercial, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_gerentecomercial, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_gerentecomercial, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_gerentecomercial, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/GERENTECOMERCIAL.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#ENTREGASVN
@login_required
def entregasvn(request, company=None):

    usuarios_entregasvn = Usuario.objects.filter(area='ENTREGAS VN')
    
    usuarios_entregasvn = usuarios_entregasvn.filter(area='ENTREGAS VN', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_entregasvn = Usuario.objects.filter(area='ENTREGAS VN')
            usuarios_entregasvn = usuarios_entregasvn.filter(area='ENTREGAS VN', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_entregasvn = Usuario.objects.filter(area='ENTREGAS VN')
            usuarios_entregasvn = usuarios_entregasvn.filter(area='ENTREGAS VN', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_entregasvn = Usuario.objects.filter(area='ENTREGAS VN')
            usuarios_entregasvn = usuarios_entregasvn.filter(area='ENTREGAS VN', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_entregasvn = Usuario.objects.filter(area='ENTREGAS VN')
            usuarios_entregasvn = usuarios_entregasvn.filter(area='ENTREGAS VN', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_entregasvn = Usuario.objects.filter(area='ENTREGAS VN')
            usuarios_entregasvn = usuarios_entregasvn.filter(area='ENTREGAS VN', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])

    telefonos_entregasvn = []
    for usuario in usuarios_entregasvn:
        telefonos_usuario_entregasvn = usuario.telefono_set.all()
        telefonos_entregasvn.extend(telefonos_usuario_entregasvn)

    impresoras_entregasvn = []
    for usuario1 in usuarios_entregasvn:
        impresoras_usuario_entregasvn = usuario1.impresora_set.all()
        impresoras_entregasvn.extend(impresoras_usuario_entregasvn)

    notebooks_entregasvn = []
    for usuario2 in usuarios_entregasvn:
        notebooks_usuario_entregasvn = usuario2.notebook_set.all()
        notebooks_entregasvn.extend(notebooks_usuario_entregasvn)

    pcs_entregasvn = []
    for usuario3 in usuarios_entregasvn:
        pcs_usuario_entregasvn = usuario3.pc_set.all()
        pcs_entregasvn.extend(pcs_usuario_entregasvn)
        
    activos_entregasvn = []
    for usuario4 in usuarios_entregasvn:
        activos_usuario_entregasvn = usuario4.activoinfraestructura_set.all()
        activos_entregasvn.extend(activos_usuario_entregasvn)
        
    pctotal = len(pcs_entregasvn)
    teltotal = len(telefonos_entregasvn)
    imptotal = len(impresoras_entregasvn)
    notetotal = len(notebooks_entregasvn)
    acttotal = len(activos_entregasvn)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_entregasvn, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_entregasvn, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_entregasvn, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_entregasvn, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_entregasvn, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/ENTREGASVN.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'acttotal':acttotal,'notetotal':notetotal,'imptotal':imptotal,'teltotal':teltotal,'pctotal':pctotal,'total':total})

#CONTACTC
@login_required
def contactc(request, company=None):

    usuarios_contactc = Usuario.objects.filter(area='CONTACT C')
    
    usuarios_contactc = usuarios_contactc.filter(area='CONTACT C', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_contactc = Usuario.objects.filter(area='CONTACT C')
            usuarios_contactc = usuarios_contactc.filter(area='CONTACT C', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_contactc = Usuario.objects.filter(area='CONTACT C')
            usuarios_contactc = usuarios_contactc.filter(area='CONTACT C', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_contactc = Usuario.objects.filter(area='CONTACT C')
            usuarios_contactc = usuarios_contactc.filter(area='CONTACT C', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_contactc = Usuario.objects.filter(area='CONTACT C')
            usuarios_contactc = usuarios_contactc.filter(area='CONTACT C', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_contactc = Usuario.objects.filter(area='CONTACT C')
            usuarios_contactc = usuarios_contactc.filter(area='CONTACT C', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_contactc = []
    for usuario in usuarios_contactc:
        telefonos_usuario_contactc = usuario.telefono_set.all()
        telefonos_contactc.extend(telefonos_usuario_contactc)

    impresoras_contactc = []
    for usuario1 in usuarios_contactc:
        impresoras_usuario_contactc = usuario1.impresora_set.all()
        impresoras_contactc.extend(impresoras_usuario_contactc)

    notebooks_contactc = []
    for usuario2 in usuarios_contactc:
        notebooks_usuario_contactc = usuario2.notebook_set.all()
        notebooks_contactc.extend(notebooks_usuario_contactc)

    pcs_contactc = []
    for usuario3 in usuarios_contactc:
        pcs_usuario_contactc = usuario3.pc_set.all()
        pcs_contactc.extend(pcs_usuario_contactc)

    activos_contactc = []
    for usuario4 in usuarios_contactc:
        activos_usuario_contactc = usuario4.activoinfraestructura_set.all()
        activos_contactc.extend(activos_usuario_contactc)
        
    pctotal = len(pcs_contactc)
    teltotal = len(telefonos_contactc)
    imptotal = len(impresoras_contactc)
    notetotal = len(notebooks_contactc)
    acttotal = len(activos_contactc)

    total = pctotal + teltotal + imptotal + notetotal + acttotal


    paginator = Paginator(telefonos_contactc, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_contactc, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_contactc, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_contactc, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)
    
    paginator = Paginator(activos_contactc, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'area/CONTACTC.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#SERVICIOSPV
@login_required
def serviciospv(request, company=None):

    usuarios_serviciospv = Usuario.objects.filter(Q(area='SERVICIOS PV') | Q(area='SERVICIOS_NEW') | Q(area='REPUESTOS_NEW'))
    
    usuarios_serviciospv = usuarios_serviciospv.filter(Q(area='SERVICIOS PV') | Q(area='SERVICIOS_NEW') | Q(area='REPUESTOS_NEW'), lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_serviciospv = Usuario.objects.filter(area='SERVICIOS PV')
            usuarios_serviciospv = usuarios_serviciospv.filter(area='SERVICIOS PV', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_serviciospv = Usuario.objects.filter(Q(area='SERVICIOS PV') | Q(area='SERVICIOS_NEW') | Q(area='REPUESTOS_NEW'))
            usuarios_serviciospv = usuarios_serviciospv.filter(Q(area='SERVICIOS PV', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL']) | Q(area='SERVICIOS_NEW', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL']) | Q(area='REPUESTOS_NEW', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL']))
        elif company == 'CITROEN':
            usuarios_serviciospv = Usuario.objects.filter(area='SERVICIOS PV')
            usuarios_serviciospv = usuarios_serviciospv.filter(area='SERVICIOS PV', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_serviciospv = Usuario.objects.filter(area='SERVICIOS PV')
            usuarios_serviciospv = usuarios_serviciospv.filter(area='SERVICIOS PV', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_serviciospv = Usuario.objects.filter(area='SERVICIOS PV')
            usuarios_serviciospv = usuarios_serviciospv.filter(area='SERVICIOS PV', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_serviciospv = []
    for usuario in usuarios_serviciospv:
        telefonos_usuario_serviciospv = usuario.telefono_set.all()
        telefonos_serviciospv.extend(telefonos_usuario_serviciospv)

    impresoras_serviciospv = []
    for usuario1 in usuarios_serviciospv:
        impresoras_usuario_serviciospv = usuario1.impresora_set.all()
        impresoras_serviciospv.extend(impresoras_usuario_serviciospv)

    notebooks_serviciospv = []
    for usuario2 in usuarios_serviciospv:
        notebooks_usuario_serviciospv = usuario2.notebook_set.all()
        notebooks_serviciospv.extend(notebooks_usuario_serviciospv)

    pcs_serviciospv = []
    for usuario3 in usuarios_serviciospv:
        pcs_usuario_serviciospv = usuario3.pc_set.all()
        pcs_serviciospv.extend(pcs_usuario_serviciospv)

    activos_serviciospv = []
    for usuario4 in usuarios_serviciospv:
        activos_usuario_serviciospv = usuario4.activoinfraestructura_set.all()
        activos_serviciospv.extend(activos_usuario_serviciospv)

    pctotal = len(pcs_serviciospv)
    teltotal = len(telefonos_serviciospv)
    imptotal = len(impresoras_serviciospv)
    notetotal = len(notebooks_serviciospv)
    acttotal = len(activos_serviciospv)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_serviciospv, 500)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_serviciospv, 500)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_serviciospv, 500)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_serviciospv, 500)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_serviciospv, 500)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/SERVICIOSPV.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#ADMVENTASPLANES
@login_required
def admventasplanes(request, company=None):

    usuarios_admventasplanes = Usuario.objects.filter(area='ADM VENTAS PLANES')
    
    usuarios_admventasplanes = usuarios_admventasplanes.filter(area='ADM VENTAS PLANES', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_admventasplanes = Usuario.objects.filter(area='ADM VENTAS PLANES')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='ADM VENTAS PLANES', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_admventasplanes = Usuario.objects.filter(area='ADM VENTAS PLANES')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='ADM VENTAS PLANES', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_admventasplanes = Usuario.objects.filter(area='ADM VENTAS PLANES')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='ADM VENTAS PLANES', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_admventasplanes = Usuario.objects.filter(area='ADM VENTAS PLANES')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='ADM VENTAS PLANES', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_admventasplanes = Usuario.objects.filter(area='ADM VENTAS PLANES')
            usuarios_admventasplanes = usuarios_admventasplanes.filter(area='ADM VENTAS PLANES', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_admventasplanes = []
    for usuario in usuarios_admventasplanes:
        telefonos_usuario_admventasplanes = usuario.telefono_set.all()
        telefonos_admventasplanes.extend(telefonos_usuario_admventasplanes)

    impresoras_admventasplanes = []
    for usuario1 in usuarios_admventasplanes:
        impresoras_usuario_admventasplanes = usuario1.impresora_set.all()
        impresoras_admventasplanes.extend(impresoras_usuario_admventasplanes)

    notebooks_admventasplanes = []
    for usuario2 in usuarios_admventasplanes:
        notebooks_usuario_admventasplanes = usuario2.notebook_set.all()
        notebooks_admventasplanes.extend(notebooks_usuario_admventasplanes)

    pcs_admventasplanes = []
    for usuario3 in usuarios_admventasplanes:
        pcs_usuario_admventasplanes = usuario3.pc_set.all()
        pcs_admventasplanes.extend(pcs_usuario_admventasplanes)
        
    activos_admventasplanes = []
    for usuario4 in usuarios_admventasplanes:
        activos_usuario_admventasplanes = usuario4.activoinfraestructura_set.all()
        activos_admventasplanes.extend(activos_usuario_admventasplanes)
        
    pctotal = len(pcs_admventasplanes)
    teltotal = len(telefonos_admventasplanes)
    imptotal = len(impresoras_admventasplanes)
    notetotal = len(notebooks_admventasplanes)
    acttotal = len(activos_admventasplanes)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_admventasplanes, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/ADMVENTASPLANES.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#VENTADIRECTA0KM
@login_required
def ventadirecta0km(request, company=None):

    usuarios_ventadirecta0km = Usuario.objects.filter(area='VENTA DIRECTA 0KM')
    
    usuarios_ventadirecta0km = usuarios_ventadirecta0km.filter(area='VENTA DIRECTA 0KM', lab_lpg=company)
    
    if company:
        if company == 'FIAT':
            usuarios_ventadirecta0km = Usuario.objects.filter(area='VENTA DIRECTA 0KM')
            usuarios_ventadirecta0km = usuarios_ventadirecta0km.filter(area='VENTA DIRECTA 0KM', lab_lpg__in=['VOLANT URQUIZA','VOLANT CENTRAL'])
        elif company == 'PEUGEOT':
            usuarios_ventadirecta0km = Usuario.objects.filter(area='VENTA DIRECTA 0KM')
            usuarios_ventadirecta0km = usuarios_ventadirecta0km.filter(area='VENTA DIRECTA 0KM', lab_lpg__in=['PEUGEOT CORDOBA','AVENUE CORDOBA','ADMINISTRACION CENTRAL'])
        elif company == 'CITROEN':
            usuarios_ventadirecta0km = Usuario.objects.filter(area='VENTA DIRECTA 0KM')
            usuarios_ventadirecta0km = usuarios_ventadirecta0km.filter(area='VENTA DIRECTA 0KM', lab_lpg__in=['IQSA CORDOBA','AUTOROUTE'])
        elif company == 'BMW':
            usuarios_ventadirecta0km = Usuario.objects.filter(area='VENTA DIRECTA 0KM')
            usuarios_ventadirecta0km = usuarios_ventadirecta0km.filter(area='VENTA DIRECTA 0KM', lab_lpg__in=['AMSA MOTORRAD','AMSA MINI','AMSA BMW'])
        elif company == 'CHEVROLET':
            usuarios_ventadirecta0km = Usuario.objects.filter(area='VENTA DIRECTA 0KM')
            usuarios_ventadirecta0km = usuarios_ventadirecta0km.filter(area='VENTA DIRECTA 0KM', lab_lpg__in=['CHEVENT VENADO TUERTO','CHEVENT SALADILLO'])


    telefonos_ventadirecta0km = []
    for usuario in usuarios_ventadirecta0km:
        telefonos_usuario_ventadirecta0km = usuario.telefono_set.all()
        telefonos_ventadirecta0km.extend(telefonos_usuario_ventadirecta0km)

    impresoras_ventadirecta0km = []
    for usuario1 in usuarios_ventadirecta0km:
        impresoras_usuario_ventadirecta0km = usuario1.impresora_set.all()
        impresoras_ventadirecta0km.extend(impresoras_usuario_ventadirecta0km)

    notebooks_ventadirecta0km = []
    for usuario2 in usuarios_ventadirecta0km:
        notebooks_usuario_ventadirecta0km = usuario2.notebook_set.all()
        notebooks_ventadirecta0km.extend(notebooks_usuario_ventadirecta0km)

    pcs_ventadirecta0km = []
    for usuario3 in usuarios_ventadirecta0km:
        pcs_usuario_ventadirecta0km = usuario3.pc_set.all()
        pcs_ventadirecta0km.extend(pcs_usuario_ventadirecta0km)

    activos_ventadirecta0km = []
    for usuario4 in usuarios_ventadirecta0km:
        activos_usuario_ventadirecta0km = usuario4.activoinfraestructura_set.all()
        activos_ventadirecta0km.extend(activos_usuario_ventadirecta0km)

    pctotal = len(pcs_ventadirecta0km)
    teltotal = len(telefonos_ventadirecta0km)
    imptotal = len(impresoras_ventadirecta0km)
    notetotal = len(notebooks_ventadirecta0km)
    acttotal = len(activos_ventadirecta0km)

    total = pctotal + teltotal + imptotal + notetotal + acttotal

    paginator = Paginator(telefonos_ventadirecta0km, 200)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_ventadirecta0km, 200)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_ventadirecta0km, 200)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_ventadirecta0km, 200)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_ventadirecta0km, 200)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'area/VENTADIRECTA0KM.html', {'page_obj': page_obj, 'page_obj1': page_obj1, 'page_obj2': page_obj2, 'page_obj3': page_obj3,'page_obj4':page_obj4,'pctotal':pctotal,'teltotal':teltotal,'imptotal':imptotal,'notetotal':notetotal,'acttotal':acttotal,'total':total})

#MARCAFIAT
@login_required
def fiat(request):

    usuarios_lavadero = Usuario.objects.filter(razon_social='VOLANT S.A.')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/FIAT.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})



#MARCABMW
@login_required
def bmw(request):

    usuarios_lavadero = Usuario.objects.filter(Q(razon_social='AUTO MUNICH S.A.') &(Q(lab_lpg='AMSA MOTORRAD') | Q(lab_lpg='AMSA BMW') | Q(lab_lpg='AMSA MINI')) &~Q(lab_lpg__in=['ADMINISTRACION CENTRAL', 'AVENUE DS'])).order_by('nombre_apellido')
  
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/BMW.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})


#MARCACHEVROLET
@login_required
def chevrolet(request):


    usuarios_lavadero = Usuario.objects.filter(
        Q(razon_social='CHEVENT S.A.') | (Q(lab_lpg='CHEVENT VENADO TUERTO') | Q(lab_lpg='CHEVENT SALADILLO')))

    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        
    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'marca/CHEVROLET.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

#MARCACITROEN
@login_required
def citroen(request):

    usuarios_lavadero = Usuario.objects.filter(Q(razon_social='IQSA S.A.') | Q(razon_social='AUTOROUTE S.A.'))
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'marca/CITROEN.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

#MARCADS
@login_required
def ds(request):

    usuarios_lavadero = Usuario.objects.filter(Q(razon_social='AVENUE S.A.') | Q(lab_lpg='AVENUE DS')).exclude(lab_lpg__in=['AVENUE ROSARIO', 'AVENUE CORDOBA', 'ADMINISTRACION CENTRAL','CHEVENT VENADO TUERTO','PEUGEOT CORDOBA'])
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)
   
    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)


    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/DS.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

#MARCARENAULT
@login_required
def renault(request):
    
    usuarios_sistema_helpdesk = Usuario.objects.filter(Q(razon_social='AILES S.A.') & Q(lab_lpg='AILES CABA'))
    
    telefonos_sistemas = []
    for usuario in usuarios_sistema_helpdesk:
        telefonos_usuario_sistemas = usuario.telefono_set.all()
        telefonos_sistemas.extend(telefonos_usuario_sistemas)

    telefonos_sistemas.sort(key=lambda tel: tel.usuario.nombre_apellido)
    
    impresoras_sistemas = []
    for usuario1 in usuarios_sistema_helpdesk:
        impresoras_usuario_sistemas = usuario1.impresora_set.all()
        impresoras_sistemas.extend(impresoras_usuario_sistemas)

    impresoras_sistemas.sort(key=lambda imp: imp.usuario.nombre_apellido)
    
    notebooks_sistemas = []
    for usuario2 in usuarios_sistema_helpdesk:
        notebooks_usuario_sistemas = usuario2.notebook_set.all()
        notebooks_sistemas.extend(notebooks_usuario_sistemas)

    notebooks_sistemas.sort(key=lambda note: note.usuario.nombre_apellido)
    
    pcs_sistemas = []
    for usuario3 in usuarios_sistema_helpdesk:
        pcs_usuario_sistemas = usuario3.pc_set.all()
        pcs_sistemas.extend(pcs_usuario_sistemas)

    pcs_sistemas.sort(key=lambda pc: pc.usuario.nombre_apellido)

    
    activos_sistemas = []
    for usuario4 in usuarios_sistema_helpdesk:
        activos_usuario_sistemas = usuario4.activoinfraestructura_set.all()
        activos_sistemas.extend(activos_usuario_sistemas)

    activos_sistemas.sort(key=lambda act: act.usuario.nombre_apellido)

    
    telefonoscont = len(telefonos_sistemas)
    notebookscont = len(notebooks_sistemas)
    pcscont = len(pcs_sistemas)
    impresorascont = len(impresoras_sistemas)
    activoscont = len(activos_sistemas)
    
    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
    
    paginator = Paginator(telefonos_sistemas, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    paginator = Paginator(impresoras_sistemas, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_sistemas, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_sistemas, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_sistemas, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render (request, 'marca/RENAULT.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

#MARCAPEUGEOT
@login_required
def peugeot(request):

    usuarios_lavadero = Usuario.objects.filter(Q(lab_lpg='ADMINISTRACION CENTRAL') | Q(lab_lpg='AVENUE CORDOBA')).exclude(lab_lpg='AVENUE DS')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/PEUGEOT.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def iqsa(request):

    usuarios_lavadero = Usuario.objects.filter(razon_social='IQSA S.A.').exclude(Q(lab_lpg='ADMINISTRACION CENTRAL') | Q(lab_lpg='AVENUE CORDOBA'))
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/IQSA.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def autoroute(request):

    usuarios_lavadero = Usuario.objects.filter(Q(razon_social='AUTOROUTE S.A.') | Q(lab_lpg='AUTOROUTE'))
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        
    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'marca/AUTOROUTE.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def avenuecba(request):

    usuarios_lavadero = Usuario.objects.filter(Q(lab_lpg='AVENUE CORDOBA')).exclude(Q(lugar_trab='SANTA FE') | Q(lugar_trab='CAPITAL FEDERAL') | Q(lugar_trab='BUENOS AIRES'))
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/AVENUECBA.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})


@login_required
def avenuero(request):

    usuarios_lavadero = Usuario.objects.filter(lab_lpg='AVENUE ROSARIO')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/AVENUERO.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def centralavenuecba(request):

    usuarios_lavadero = Usuario.objects.filter(lab_lpg='ADMINISTRACION CENTRAL').exclude(lugar_trab='SANTA FE')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/CENTRALAVENUECBA.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def chevents(request):
    usuarios_lavadero = Usuario.objects.filter(lab_lpg='CHEVENT SALADILLO')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'marca/CHEVENTS.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def cheventvt(request):

    usuarios_lavadero = Usuario.objects.filter(lab_lpg='CHEVENT VENADO TUERTO')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/CHEVENTVT.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def amsabmw(request):

    usuarios_lavadero = Usuario.objects.filter(razon_social='AUTO MUNICH S.A.', lab_lpg='AMSA BMW')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)



    return render(request, 'marca/AMSABMW.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def amsamini(request):

    usuarios_lavadero = Usuario.objects.filter(razon_social='AUTO MUNICH S.A.', lab_lpg='AMSA MINI')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)



    return render(request, 'marca/AMSAMINI.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def amsamotorrad(request):

    usuarios_lavadero = Usuario.objects.filter(razon_social='AUTO MUNICH S.A.', lab_lpg='AMSA MOTORRAD')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)



    return render(request, 'marca/AMSAMOTORRAD.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def fiatcentral(request):

    usuarios_lavadero = Usuario.objects.filter(razon_social='VOLANT S.A.', lab_lpg='VOLANT CENTRAL')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        

    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)


    return render(request, 'marca/FIATCENTRAL.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def fiaturquiza(request):

    usuarios_lavadero = Usuario.objects.filter(razon_social='VOLANT S.A.', lab_lpg='VOLANT URQUIZA')
    
    telefonos_lavadero = []
    for usuario in usuarios_lavadero:
        telefonos_usuario_lavadero = usuario.telefono_set.all()
        telefonos_lavadero.extend(telefonos_usuario_lavadero)

    telefonos_lavadero.sort(key=lambda tel: tel.usuario.nombre_apellido)

    impresoras_lavadero = []
    for usuario1 in usuarios_lavadero:
        impresoras_usuario_lavadero = usuario1.impresora_set.all()
        impresoras_lavadero.extend(impresoras_usuario_lavadero)

    impresoras_lavadero.sort(key=lambda imp: imp.usuario.nombre_apellido)

    notebooks_lavadero = []
    for usuario2 in usuarios_lavadero:
        notebooks_usuario_lavadero = usuario2.notebook_set.all()
        notebooks_lavadero.extend(notebooks_usuario_lavadero)

    notebooks_lavadero.sort(key=lambda note: note.usuario.nombre_apellido)

    pcs_lavadero = []
    for usuario3 in usuarios_lavadero:
        pcs_usuario_lavadero = usuario3.pc_set.all()
        pcs_lavadero.extend(pcs_usuario_lavadero)

    pcs_lavadero.sort(key=lambda pc: pc.usuario.nombre_apellido)

    activos_lavadero = []
    for usuario4 in usuarios_lavadero:
        activos_usuario_lavadero = usuario4.activoinfraestructura_set.all()
        activos_lavadero.extend(activos_usuario_lavadero)

    activos_lavadero.sort(key=lambda act: act.usuario.nombre_apellido)

    telefonoscont = len(telefonos_lavadero)
    notebookscont = len(notebooks_lavadero)
    pcscont = len(pcs_lavadero)
    impresorascont = len(impresoras_lavadero)
    activoscont = len(activos_lavadero)

    total = telefonoscont + notebookscont + pcscont + impresorascont + activoscont
        
    paginator = Paginator(telefonos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator = Paginator(impresoras_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj1 = paginator.get_page(page_number)

    paginator = Paginator(notebooks_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number)

    paginator = Paginator(pcs_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj3 = paginator.get_page(page_number)

    paginator = Paginator(activos_lavadero, 250)
    page_number = request.GET.get('page')
    page_obj4 = paginator.get_page(page_number)

    return render(request, 'marca/FIATURQUIZA.html',{'page_obj':page_obj,'page_obj1':page_obj1,'page_obj2':page_obj2,'page_obj3':page_obj3,'page_obj4':page_obj4,'telefonoscont':telefonoscont,'notebookscont':notebookscont,'pcscont':pcscont,'impresorascont':impresorascont,'activoscont':activoscont,'total':total})

@login_required
def verificar_notebook(request, user_id):
    usuario_notebook = Notebook.objects.filter(usuario_id=user_id).exists()
    response = {
        'tieneNotebook': usuario_notebook
    }
    return JsonResponse(response)

#ABONOCREAR
@login_required
def abono_crear(request):
    if request.method == 'POST':
        form = AbonoCelForm(request.POST)
        if form.is_valid():
            abonos = form.save(commit=False)
            abonos.save()
            print(abonos)
            return redirect('inicio')
    else:
        form = AbonoCelForm()
    return render(request, 'sistema/abono_crear.html', {'form': form})

#ABONOCELLISTAR
@login_required
def abono_listar(request):

    contador = AbonoCelular.objects.all().count()
    if request.method == 'POST':
        precios_actualizados = request.POST.getlist('precio')

        for precio_actualizado, abono in zip(precios_actualizados, AbonoCelular.objects.all()):
            precio_actualizado = precio_actualizado.replace(',', '.')
            abono.precio = precio_actualizado
            abono.save()

        return redirect('inicio')

    abonos = AbonoCelular.objects.all()
    paginator_usuarios = Paginator(abonos, 200)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    return render(request, 'sistema/abono_listar.html', {'page_obj_usuarios': page_obj_usuarios, 'contador':contador})

#CALCULAR ABONOS CELULAR
@login_required
def calcularabonocel(request):


    tel_iqsa = Telefono.objects.filter(Q(lablpg='IQSA CORDOBA') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_avenue_rosario = Telefono.objects.filter(Q(lablpg='AVENUE ROSARIO') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_autoroute = Telefono.objects.filter(Q(lablpg='AUTOROUTE') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_chevent_saladillo = Telefono.objects.filter(Q(lablpg='CHEVENT SALADILLO') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_peugeot_cordoba = Telefono.objects.filter(Q(lablpg='PEUGEOT CORDOBA') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_adm_central_cba = Telefono.objects.filter(Q(lablpg='ADMINISTRACION CENTRAL') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_avenue_cordoba = Telefono.objects.filter(Q(lablpg='AVENUE CORDOBA') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    cant_avenue_cordoba = tel_peugeot_cordoba + tel_adm_central_cba + tel_avenue_cordoba
    tel_volant_urquiza = Telefono.objects.filter(Q(lablpg='VOLANT URQUIZA') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_amsa_bmw = Telefono.objects.filter(Q(lablpg='AMSA BMW') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_chevent_venado_tuerto = Telefono.objects.filter(Q(lablpg='CHEVENT VENADO TUERTO') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_volant_central = Telefono.objects.filter(Q(lablpg='VOLANT CENTRAL') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_ds = Telefono.objects.filter(Q(lablpg='AVENUE DS') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_avenue_rosario = Telefono.objects.filter(Q(lablpg='AVENUE ROSARIO') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()
    tel_ailes_caba = Telefono.objects.filter(Q(lablpg='AILES CABA') & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').count()


    buscar_query = request.GET.get('nombre')
    buscar_query1 = request.GET.get('rs')
    buscar_query2 = request.GET.get('lablpg')

    
    telefonos = Telefono.objects.all().order_by('usuario__nombre_apellido')
    total = len(telefonos)
    abono = AbonoCelular.objects.all()
    totalpre_con_separadores = 0
    totalpre = 0

    
    if request.method == 'GET':
        nombres = request.GET.get('nombre', '')
        rs = request.GET.get('rs', '')
        lablpg = request.GET.get('lablpg','')
        context = {
            'nombres': Telefono.PLAN_CHOICES,
            'rs': Telefono.RS_CHOICES,
            'lablpg': Telefono.LABPLG_CHOICES
        }
    
    if buscar_query:
        telefonos = Telefono.objects.filter(Q(plan__exact=buscar_query)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        total = len(telefonos)
        abono = AbonoCelular.objects.filter(Q(nombre__icontains=buscar_query)).values_list('precio').first()[0]
        totalpre = abono * len(telefonos)
        totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
        print('AAA')
        print(telefonos)
        if buscar_query1:
            if total == 0:
                print('no hay nada en buscar query 1')
            else:
                telefonos = Telefono.objects.filter(Q(rs__exact=buscar_query1) & Q(plan__exact=buscar_query)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
                total = len(telefonos)
                abono = AbonoCelular.objects.filter(Q(nombre__exact=buscar_query)).values_list('precio').first()[0]
                totalpre = abono * len(telefonos)
                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                print(telefonos)
                print('entro aca1')
        elif buscar_query2:
            if total == 0:
                print('no hay nada en buscar query2')
            else:
                telefonos = Telefono.objects.filter(Q(lablpg__exact=buscar_query2)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
                total = len(telefonos)
                abono = AbonoCelular.objects.filter(Q(nombre__exact=buscar_query)).values_list('precio').first()[0]
                totalpre = abono * len(telefonos)
                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                print('entro aca2')
    elif buscar_query1:
        telefonos = Telefono.objects.filter(Q(rs__exact=buscar_query1)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').order_by('usuario__nombre_apellido')
        totalpre = 0

        for telefono in telefonos:
            plan_telefono = telefono.plan
            abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
            for abono_celular in abonos_celular:
                totalpre += abono_celular.precio

        totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
        
        print(totalpre)
        total = len(telefonos)

        if buscar_query2:
            if total == 0:
                print('no hay nada en buscar query 1')
            else:
                telefonos = Telefono.objects.filter(Q(rs__exact=buscar_query1) & Q(lablpg__exact=buscar_query2)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
                totalpre = 0
                total = len(telefonos)

                for telefono in telefonos:
                    plan_telefono = telefono.plan
                    abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
                    for abono_celular in abonos_celular:
                        totalpre += abono_celular.precio
                
                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                total = len(telefonos)
                
    elif buscar_query2:
        if buscar_query2 == 'AVENUE CORDOBA':
            lablpg_values = ['AVENUE CORDOBA', 'PEUGEOT CORDOBA', 'ADMINISTRACION CENTRAL']
            telefonos = Telefono.objects.filter(Q(lablpg__in=lablpg_values) & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

            totalpre = 0
            for telefono in telefonos:
                plan_telefono = telefono.plan
                abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
                for abono_celular in abonos_celular:
                    totalpre += abono_celular.precio

                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                total = len(telefonos)
        else:
            telefonos = Telefono.objects.filter(Q(lablpg__exact=buscar_query2) & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
            totalpre = 0

            for telefono in telefonos:
                plan_telefono = telefono.plan
                abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
                for abono_celular in abonos_celular:
                    totalpre += abono_celular.precio

                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                total = len(telefonos)
    
    paginator_usuarios = Paginator(telefonos, 2000)
    page_number_usuarios = request.GET.get('page')
    page_obj_telefonos = paginator_usuarios.get_page(page_number_usuarios)
    
    return render(request, 'sistema/calcularabonocel.html',{'tel_ailes_caba':tel_ailes_caba,'tel_avenue_rosario':tel_avenue_rosario,'tel_ds':tel_ds,'tel_volant_central':tel_volant_central,'tel_chevent_venado_tuerto':tel_chevent_venado_tuerto,'tel_amsa_bmw':tel_amsa_bmw,'tel_volant_urquiza':tel_volant_urquiza,'tel_avenue_cordoba':cant_avenue_cordoba,'tel_chevent_saladillo':tel_chevent_saladillo,'tel_autoroute':tel_autoroute,'tel_avenue_rosario':tel_avenue_rosario,'tel_iqsa':tel_iqsa,'telefonos':telefonos,'page_obj_telefonos':page_obj_telefonos,'total':total,'totalpre':totalpre,'abono':abono,'totalpre_con_separadores':totalpre_con_separadores, **context})

def descargar_excel_abono(request):

    buscar_query = request.GET.get('nombre')
    buscar_query1 = request.GET.get('rs')
    buscar_query2 = request.GET.get('lablpg')
    
    telefonos = Telefono.objects.all().order_by('usuario__nombre_apellido')
    total = len(telefonos)
    abono = AbonoCelular.objects.all()
    totalpre_con_separadores = 0
    totalpre = 0
    
    if request.method == 'GET':
        nombres = request.GET.get('nombre', '')
        rs = request.GET.get('rs', '')
        lablpg = request.GET.get('lablpg','')
        context = {
            'nombres': Telefono.PLAN_CHOICES,
            'rs': Telefono.RS_CHOICES,
            'lablpg': Telefono.LABPLG_CHOICES
        }
    
    if buscar_query:
        telefonos = Telefono.objects.filter(Q(plan__exact=buscar_query)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
        total = len(telefonos)
        abono = AbonoCelular.objects.filter(Q(nombre__icontains=buscar_query)).values_list('precio').first()[0]
        totalpre = abono * len(telefonos)
        totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
        print('AAA')
        print(telefonos)
        if buscar_query1:
            if total == 0:
                print('no hay nada en buscar query 1')
            else:
                telefonos = Telefono.objects.filter(Q(rs__exact=buscar_query1) & Q(plan__exact=buscar_query)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
                total = len(telefonos)
                abono = AbonoCelular.objects.filter(Q(nombre__exact=buscar_query)).values_list('precio').first()[0]
                totalpre = abono * len(telefonos)
                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                print(telefonos)
                print('entro aca1')
        elif buscar_query2:
            if total == 0:
                print('no hay nada en buscar query2')
            else:
                telefonos = Telefono.objects.filter(Q(lablpg__exact=buscar_query2)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
                total = len(telefonos)
                abono = AbonoCelular.objects.filter(Q(nombre__exact=buscar_query)).values_list('precio').first()[0]
                totalpre = abono * len(telefonos)
                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                print('entro aca2')
    elif buscar_query1:
        telefonos = Telefono.objects.filter(Q(rs__exact=buscar_query1)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').order_by('usuario__nombre_apellido')
        totalpre = 0

        for telefono in telefonos:
            plan_telefono = telefono.plan
            abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
            for abono_celular in abonos_celular:
                totalpre += abono_celular.precio

        totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
        
        print(totalpre)
        total = len(telefonos)

        if buscar_query2:
            if total == 0:
                print('no hay nada en buscar query 1')
            else:
                telefonos = Telefono.objects.filter(Q(rs__exact=buscar_query1) & Q(lablpg__exact=buscar_query2)).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
                totalpre = 0
                total = len(telefonos)

                for telefono in telefonos:
                    plan_telefono = telefono.plan
                    abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
                    for abono_celular in abonos_celular:
                        totalpre += abono_celular.precio
                
                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                total = len(telefonos)
                
    elif buscar_query2:
        if buscar_query2 == 'PEUGEOT CORDOBA':
            lablpg_values = ['AVENUE CORDOBA', 'PEUGEOT CORDOBA', 'ADMINISTRACION CENTRAL']
            telefonos = Telefono.objects.filter(Q(lablpg__in=lablpg_values) & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

            totalpre = 0
            for telefono in telefonos:
                plan_telefono = telefono.plan
                abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
                for abono_celular in abonos_celular:
                    totalpre += abono_celular.precio

                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                total = len(telefonos)
        else:
            telefonos = Telefono.objects.filter(Q(lablpg__exact=buscar_query2) & ~Q(usuario__nombre_apellido='Sin_asignar')).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')
            totalpre = 0

            for telefono in telefonos:
                plan_telefono = telefono.plan
                abonos_celular = AbonoCelular.objects.filter(nombre=plan_telefono)
                for abono_celular in abonos_celular:
                    totalpre += abono_celular.precio

                totalpre_con_separadores = "{:,.2f}".format(totalpre).replace(',', 'temp').replace('.', ',').replace('temp', '.')
                total = len(telefonos)

    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    

    # Agregar encabezados
    headers = ['Usuario','Area', 'Numero', 'Modelo', 'Marca', 'Accesorio', 'Plan', 'Empresa Abono', 'RS', 'LabLPG','Total']
    sheet.append(headers)

    # Agregar datos a las filas
    for telefono in telefonos:
        row = [
            telefono.usuario.nombre_apellido,
            telefono.usuario.area,
            telefono.numero,
            telefono.modelo,
            telefono.marca,
            telefono.accesorio,
            telefono.plan,
            telefono.empresa_abono,
            telefono.rs,
            telefono.lablpg
        ]
        sheet.append(row)

    # Agregar la fila de totalpre_con_separadores al final
    sheet.append(['', '', '', '', '', '', '', '', '', '', totalpre_con_separadores])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=telefonos.xlsx'
    workbook.save(response)
    return response


#CREAR LICENCIAS MICROSOFT
@login_required
def usuario_licencias(request, usuario_id):

    usuario = get_object_or_404(Usuario, pk=usuario_id)
    licencias = usuario.licencia_set.all()

    if request.method == 'POST':
        form = LicenciaForm(request.POST)
        if form.is_valid():
            nueva_licencia = form.save(commit=False)

            if not Licencia.objects.filter(usuario=usuario, licenciaoffice=nueva_licencia.licenciaoffice).exists():
                nueva_licencia.usuario = usuario
                nueva_licencia.save()

                # Crear instancia de AltaLicenciaoffice
                alta_licencia = AltaLicenciaoffice(
                    licencia=nueva_licencia,
                    nombrelicencia=nueva_licencia.get_licenciaoffice_display(),
                    usuario_alta=request.user,
                    fecha_alta=datetime.now(),
                )
                alta_licencia.save()

            else:
                form.add_error('licenciaoffice', 'Ya existe una licencia de este tipo para este usuario.')
    else:
        form = LicenciaForm()
    return render(request, 'sistema/licencia.html', {'form': form, 'usuario': usuario, 'licencias': licencias})

#ELIMINAR LICENCIA
@login_required
def eliminar_licencia(request, licencia_id):
    licencia = get_object_or_404(Licencia, pk=licencia_id)
    usuario_id = licencia.usuario_id
    licencia.delete()
    return redirect('licencia', usuario_id=usuario_id)

    path('sistema/licencia/<int:usuario_id>', views.usuario_licencias, name='licencia'),

#LISTA LICENCIAS
@login_required
def lista_licencias(request, usuario_id):
    usuario = get_object_or_404(Usuario, pk=usuario_id)
    
    return render(request, 'sistema/lista_licencias.html', {'usuario': usuario})


#ABONO LISTAR IMP
@login_required
def abono_listar_imp(request):

    contador = AbonoImpresora.objects.all().count()

    if request.method == 'POST':

        precios_actualizados = request.POST.getlist('precio')

        for precio_actualizado, abono in zip(precios_actualizados, AbonoImpresora.objects.all()):

            precio_actualizado = precio_actualizado.replace(',', '.')
            abono.precio = precio_actualizado
            abono.save()

        return redirect('inicio')

    abonos = AbonoImpresora.objects.all()

    paginator_usuarios = Paginator(abonos, 500)
    page_number_usuarios = request.GET.get('page')
    page_obj_usuarios = paginator_usuarios.get_page(page_number_usuarios)

    return render(request, 'sistema/abono_listar_imp.html', {'page_obj_usuarios': page_obj_usuarios,'contador':contador})

#LISTA LICENCIA OFF
@login_required
def lista_licencia_off(request):

    licencia = request.GET.get('licencia')
    sucursal = request.GET.get('lablpg')
    buscar_usuario = request.GET.get('buscar_usuario')

    usuarios = Usuario.objects.all()

    licencias = Licencia.objects.all().order_by('usuario__nombre_apellido')
    contador = Licencia.objects.all().count()

    precio_total = 0

    cantidad_office_e1 = Licencia.objects.filter(licenciaoffice='Office 365 E1').count()
    cantidad_office_e5 = Licencia.objects.filter(licenciaoffice='Office 365 E5').count()
    cantidad_office_e5wac = Licencia.objects.filter(licenciaoffice='Office 365 E5 sin Audioconferencia').count()
    cantidad_office_exchangeplan1 = Licencia.objects.filter(licenciaoffice='Exchange Online (plan 1)').count()
    cantidad_office_quioscoonline = Licencia.objects.filter(licenciaoffice='Quiosco de Exchange Online').count()
    
    if buscar_usuario:
        usuarios = usuarios.filter(nombre_apellido__icontains=buscar_usuario)
        licencias = Licencia.objects.filter(usuario__in=usuarios).order_by('usuario__nombre_apellido')
        contador = licencias.count()

    if licencia and sucursal:
        if sucursal == 'AVENUE CORDOBA':
            usuarios = Usuario.objects.filter(Q(lab_lpg__in=['AVENUE CORDOBA', 'ADMINISTRACION CENTRAL']))
        else:
            usuarios = Usuario.objects.filter(Q(lab_lpg=sucursal))
        licencias = Licencia.objects.filter(licenciaoffice=licencia, usuario__in=usuarios).order_by('usuario__nombre_apellido')
        contador = licencias.count()
 
    if sucursal and not licencia:
        if sucursal == 'AVENUE CORDOBA':
            usuarios = Usuario.objects.filter(Q(lab_lpg__in=['AVENUE CORDOBA', 'ADMINISTRACION CENTRAL']))
        else:
            usuarios = Usuario.objects.filter(Q(lab_lpg=sucursal))
        licencias = Licencia.objects.filter(usuario__in=usuarios).order_by('usuario__nombre_apellido')
        cantidad_office_e1 = licencias.filter(licenciaoffice='Office 365 E1').count()
        cantidad_office_e5 = licencias.filter(licenciaoffice='Office 365 E5').count()
        cantidad_office_e5wac = licencias.filter(licenciaoffice='Office 365 E5 sin Audioconferencia').count()
        cantidad_office_exchangeplan1 = licencias.filter(licenciaoffice='Exchange Online (plan 1)').count()
        cantidad_office_quioscoonline = licencias.filter(licenciaoffice='Quiosco de Exchange Online').count()
        contador = licencias.count()

    if licencia and not sucursal:
        licencias = Licencia.objects.filter(licenciaoffice=licencia).order_by('usuario__nombre_apellido')
        contador = licencias.count()

    for licencia in licencias:
        abono = AbonoOffice.objects.get(nombre=licencia.licenciaoffice)
        precio_total += abono.precio

    total_iva = (precio_total * 21) / 100 + precio_total

    licencia_nombre = next((name for value, name in Licencia.OFFICE_CHOICES if value == licencia), '')
    sucursal_nombre = next((name for value, name in Usuario.LABPLG_CHOICES if value == sucursal), '')

    total_abonos_con_separadores = "{:,.2f}".format(total_iva).replace(',', 'temp').replace('.', ',').replace('temp', '.')
    
    return render(request, 'sistema/lista_licencia_off.html', {'sucursal_nombre': sucursal_nombre,'licencias': licencias,'contador': contador,'cantidad_office_e1': cantidad_office_e1,'cantidad_office_e5':cantidad_office_e5,'cantidad_office_e5wac':cantidad_office_e5wac,'cantidad_office_exchangeplan1':cantidad_office_exchangeplan1,'cantidad_office_quioscoonline':cantidad_office_quioscoonline,'licencia_nombre': licencia_nombre,'usuarios': usuarios,'precio_total':total_abonos_con_separadores,})

#CALCULAR ABONO IMP
def calcularabonoimp(request):

    total = 0
    buscar_query1 = request.GET.get('lablpg')
    sucursal_seleccionada = buscar_query1

    imp_ailes = Impresora.objects.filter(tipo='Alquilada', lablpg='AILES CABA').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_bmw = Impresora.objects.filter(tipo='Alquilada', lablpg='AMSA BMW').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_avenue_rosario = Impresora.objects.filter(tipo='Alquilada', lablpg='AVENUE ROSARIO').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_ds = Impresora.objects.filter(tipo='Alquilada', lablpg='AVENUE DS').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_volant_central = Impresora.objects.filter(tipo='Alquilada', lablpg='VOLANT CENTRAL').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_chevent_vt = Impresora.objects.filter(tipo='Alquilada', lablpg='CHEVENT VENADO TUERTO').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_volant_urquiza = Impresora.objects.filter(tipo='Alquilada', lablpg='VOLANT URQUIZA').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_avenue_cba = Impresora.objects.filter(tipo='Alquilada', lablpg='AVENUE CORDOBA').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_autoroute = Impresora.objects.filter(tipo='Alquilada', lablpg='AUTOROUTE').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_iqsa = Impresora.objects.filter(tipo='Alquilada', lablpg='IQSA CORDOBA').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    imp_chevent_saladillo = Impresora.objects.filter(tipo='Alquilada', lablpg='CHEVENT SALADILLO').exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido').count()
    impresoras = Impresora.objects.filter(tipo='Alquilada', lablpg__exact=sucursal_seleccionada).exclude(reparabilidad='Irreparable').order_by('usuario__nombre_apellido')

    total_abonos = 0

    for impresora in impresoras:
        plan_impresora = impresora.plan
        abonos_impresora = AbonoImpresora.objects.filter(nombre=plan_impresora)
        for abono_impresora in abonos_impresora:
            total_abonos += abono_impresora.precio


    total_iva = (total_abonos * 21) / 100 + total_abonos
    total_abonos_con_separadores = "{:,.2f}".format(total_iva).replace(',', 'temp').replace('.', ',').replace('temp', '.')
    
    cant_impresoras = impresoras.count()

    paginator_impresoras = Paginator(impresoras, 2000)
    page_number_impresoras = request.GET.get('page')
    page_obj_impresoras = paginator_impresoras.get_page(page_number_impresoras)

    context = {

        'lablpg': Impresora.LABPLG_CHOICES,
        'total_abonos_con_separadores': total_abonos_con_separadores,
        'page_obj_impresoras': page_obj_impresoras,
        'cant_impresoras':cant_impresoras,
        'imp_iqsa':imp_iqsa,
        'imp_chevent_saladillo':imp_chevent_saladillo,
        'imp_autoroute':imp_autoroute,
        'imp_avenue_cba':imp_avenue_cba,
        'imp_volant_urquiza':imp_volant_urquiza,
        'imp_chevent_vt':imp_chevent_vt,
        'imp_volant_central':imp_volant_central,
        'imp_ds':imp_ds,
        'imp_avenue_rosario':imp_avenue_rosario,
        'imp_bmw':imp_bmw,
        'imp_ailes':imp_ailes,

    }

    return render(request, 'sistema/calcularabonoimp.html', context)

#ABONO OFFICES
def abono_office(request):

    contador = AbonoOffice.objects.all().count()

    if request.method == 'POST':

        precios_actualizados = request.POST.getlist('precio')

        for precio_actualizado, abono in zip(precios_actualizados, AbonoOffice.objects.all()):

            precio_actualizado = precio_actualizado.replace(',', '.')
            abono.precio = precio_actualizado
            abono.save()

        return redirect('inicio')

    abonos = AbonoOffice.objects.all()

    paginator_licencias = Paginator(abonos, 1000)
    page_number_licencias = request.GET.get('page')
    page_obj_licencias = paginator_licencias.get_page(page_number_licencias)

    return render(request,'sistema/abono_listar_off.html',{'contador':contador,'page_obj_licencias':page_obj_licencias})